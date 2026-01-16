from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import base64
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import random
import requests
from passlib.context import CryptContext
import pytz

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"

# SMS Configuration
DEFAULT_SMS_USERNAME = os.environ.get('TEXTIT_USERNAME', '942021070701')
DEFAULT_SMS_PASSWORD = os.environ.get('TEXTIT_PASSWORD', '7470')

# ============= MODELS =============
class Company(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    admin_name: str
    admin_mobile: str
    email: Optional[str] = None
    address: Optional[str] = None
    contact_number: Optional[str] = None
    short_code: Optional[str] = None  # Company short code for fingerprint device (max 20 chars)
    status: str = "pending"  # pending, active, suspended
    sms_gateway: str = "textit"
    sms_enabled: bool = False
    sms_username: Optional[str] = None
    sms_password: Optional[str] = None
    company_info_completed: bool = False
    invoicing_enabled: bool = False
    location_tracking_enabled: bool = False
    tin: Optional[str] = None  # Tax Identification Number for VAT invoices
    place_of_supply: Optional[str] = None  # Default place of supply for invoices
    branch_code: Optional[str] = None  # Branch code for invoice numbering (QQQQ in YYMMM_QQQQ_XXXXX)
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    last_login: Optional[str] = None

class CompanyCreate(BaseModel):
    name: str
    admin_name: str
    admin_mobile: str
    email: Optional[str] = None

class CompanyInfoUpdate(BaseModel):
    name: str
    address: str
    contact_number: str
    email: str

class SMSSettings(BaseModel):
    sms_gateway: str  # textit, dialog, hutch, mobitel, disabled
    sms_enabled: bool
    # Textit.biz
    sms_username: Optional[str] = None
    sms_password: Optional[str] = None
    # Dialog
    dialog_username: Optional[str] = None
    dialog_password: Optional[str] = None
    dialog_mask: Optional[str] = None  # Sender ID
    # Hutch
    hutch_client_id: Optional[str] = None
    hutch_client_secret: Optional[str] = None
    hutch_access_token: Optional[str] = None
    hutch_refresh_token: Optional[str] = None
    # Mobitel
    mobitel_app_id: Optional[str] = None
    mobitel_app_key: Optional[str] = None
    mobitel_client_id: Optional[str] = None

class User(BaseModel):
    model_config = ConfigDict(extra="allow")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: Optional[str] = None
    employee_id: Optional[str] = None
    mobile: str
    name: str
    role: str  # super_admin, admin, manager, accountant, employee, staff_member
    department: Optional[str] = None
    position: Optional[str] = None
    basic_salary: float = 0.0
    allowances: float = 0.0
    join_date: Optional[str] = None
    profile_pic: Optional[str] = None
    start_time: Optional[str] = None
    finish_time: Optional[str] = None
    fixed_salary: bool = False
    custom_start_time: Optional[str] = None
    custom_end_time: Optional[str] = None
    ot_allowed: bool = False
    sms_notifications: bool = False
    fingerprint_id: Optional[str] = None
    is_active: bool = True
    can_full_access_companies: bool = False  # For super admins: allow full edit access when viewing company portals
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UserCreate(BaseModel):
    employee_id: Optional[str] = None
    mobile: str
    name: str
    role: str
    department: Optional[str] = None
    position: Optional[str] = None
    basic_salary: float = 0.0
    allowances: float = 0.0
    join_date: str
    start_time: Optional[str] = None
    finish_time: Optional[str] = None
    fixed_salary: Optional[bool] = False
    fingerprint_id: Optional[str] = None


class BulkEmployeeParsed(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    mobile: Optional[str] = None
    role: Optional[str] = None
    position: Optional[str] = None
    department: Optional[str] = None
    join_date: Optional[str] = None
    basic_salary: Optional[float] = None
    allowances: Optional[float] = None
    start_time: Optional[str] = None
    finish_time: Optional[str] = None
    fixed_salary: Optional[bool] = False
    error: Optional[str] = None  # To track any parsing errors

class BulkEmployeeImportRequest(BaseModel):
    employees: List[dict]  # List of employee dictionaries to import

class OTPRequest(BaseModel):
    mobile: str

class OTPVerify(BaseModel):
    mobile: str
    otp: str
    login_as: Optional[str] = None

class ActivityLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    user_id: str
    user_name: str
    action: str
    details: str
    timestamp: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Increment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    employee_id: str
    employee_name: str
    old_salary: float
    new_salary: float
    increment_amount: float
    effective_from: str  # Format: "YYYY-MM"
    reason: str
    status: str = "pending"  # pending, active
    created_by: str
    created_by_name: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Advance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    employee_id: str
    employee_name: str
    amount: float
    reason: str
    repayment_months: int = 1
    status: str = "pending"  # pending, approved, rejected
    request_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).date().isoformat())
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Leave(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    employee_id: str
    employee_name: str
    leave_type: str
    from_date: str
    to_date: str
    reason: str
    status: str = "pending"  # pending, approved, rejected
    applied_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).date().isoformat())
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ============= INVOICING MODELS =============
class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    name: str
    company_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    whatsapp: Optional[str] = None
    city: Optional[str] = None
    address: Optional[str] = None
    tin: Optional[str] = None  # Tax Identification Number
    bank_name: Optional[str] = None
    bank_branch: Optional[str] = None
    bank_account_number: Optional[str] = None
    bank_account_holder_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    deleted: bool = False
    deleted_at: Optional[str] = None
    deleted_by: Optional[str] = None

class ProductCategory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    name: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    category_id: Optional[str] = None
    name: str
    description: Optional[str] = None
    price: float
    unit: str = "pcs"  # pcs, kg, hrs, etc
    stock_quantity: float = 0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    deleted: bool = False
    deleted_at: Optional[str] = None
    deleted_by: Optional[str] = None

class InvoiceItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_id: Optional[str] = None
    product_name: str
    description: Optional[str] = None
    quantity: float
    unit_price: float
    total: float
    display_amounts: bool = True  # For estimates: whether to show amounts in view

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    customer_id: str
    invoice_number: str  # Format: YYMMM_QQQQ_XXXXX (e.g., 25JAN_MAIN_00001)
    invoice_date: str
    due_date: Optional[str] = None
    date_of_delivery: Optional[str] = None  # Date when goods/services were delivered
    place_of_supply: Optional[str] = None  # Location from which delivery originates
    items: List[InvoiceItem] = []
    subtotal: float  # Net amount (exclusive of VAT)
    vat_rate: float = 18.0  # VAT rate percentage
    vat_amount: float = 0  # VAT amount calculated
    total: float  # Total including VAT
    total_in_words: Optional[str] = None  # Total amount in words
    amount_paid: float = 0
    status: str = "unpaid"  # unpaid, partial, paid
    payment_mode: Optional[str] = None  # Cash, Bank Transfer, Cheque, Credit/Debit card, Mobile Payment, Online Payment
    notes: Optional[str] = None
    created_by: str
    created_by_name: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    deleted: bool = False
    deleted_at: Optional[str] = None
    deleted_by: Optional[str] = None

class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_id: str
    amount: float
    payment_date: str
    payment_method: str  # cash, bank_transfer, cheque, card
    notes: Optional[str] = None
    created_by: str
    created_by_name: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Estimate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    customer_id: str
    estimate_number: str
    estimate_date: str
    valid_until: Optional[str] = None
    items: List[InvoiceItem] = []
    subtotal: float
    total: float
    status: str = "draft"  # draft, sent, accepted, rejected, converted
    notes: Optional[str] = None
    display_total_amounts: bool = True  # Whether to show amounts in estimate view
    created_by: str
    created_by_name: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    deleted: bool = False
    deleted_at: Optional[str] = None
    deleted_by: Optional[str] = None
    approval_status: str = "pending"  # pending, approved, rejected
    approved_by: Optional[str] = None
    approved_by_name: Optional[str] = None
    approved_at: Optional[str] = None
    rejected_by: Optional[str] = None
    rejected_by_name: Optional[str] = None
    rejected_at: Optional[str] = None


# ============= LOCATION TRACKING MODELS =============
class LocationPoint(BaseModel):
    latitude: float
    longitude: float
    timestamp: str
    accuracy: Optional[float] = None

class TrackingSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    employee_id: str
    employee_name: str
    start_time: str
    end_time: Optional[str] = None
    status: str = "active"  # active, stopped
    locations: List[LocationPoint] = []
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class LocationUpdate(BaseModel):
    session_id: str
    latitude: float
    longitude: float
    accuracy: Optional[float] = None

class AttendanceWithLocation(BaseModel):
    employee_id: Optional[str] = None
    date: str
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    status: str = "present"
    leave_type: Optional[str] = None
    latitude: float
    longitude: float
    accuracy: Optional[float] = None
    address: Optional[str] = None
    map_snapshot: Optional[str] = None  # base64 encoded image


# ============= DEVICE IMPORT MODELS =============
class DeviceImportParseRequest(BaseModel):
    file_content: str  # Raw file content
    company_id: str

class ParsedDeviceRecord(BaseModel):
    vendor_id: str
    datetime: str
    date: str
    time: str

class DeviceImportMapping(BaseModel):
    vendor_id: str
    employee_id: str
    employee_name: str

class DeviceImportRequest(BaseModel):
    company_id: str
    mappings: List[DeviceImportMapping]
    parsed_records: List[ParsedDeviceRecord]
    duplicate_action: str  # "skip" or "overwrite"


# ============= HELPER FUNCTIONS =============
def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=7)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return encoded_jwt

async def get_effective_salary(employee_id: str, company_id: str, for_month: str) -> float:
    """
    Get the effective salary for an employee for a specific month
    Considers increment history and returns the salary that was effective in that month
    
    Args:
        employee_id: Employee ID
        company_id: Company ID
        for_month: Month in format "YYYY-MM"
    
    Returns:
        Effective salary for that month
    """
    # Get employee's current/base salary
    employee = await db.users.find_one({"id": employee_id, "company_id": company_id})
    if not employee:
        return 0.0
    
    current_salary = employee.get("basic_salary", 0.0)
    
    # Get all increments for this employee that are effective on or before the target month
    increments = await db.increments.find({
        "employee_id": employee_id,
        "company_id": company_id,
        "effective_from": {"$lte": for_month}
    }).sort("effective_from", -1).to_list(length=1)
    
    # If there's an increment effective for this month or earlier, use the new salary
    if increments:
        return increments[0]["new_salary"]
    
    # Otherwise, return current salary
    return current_salary

def send_sms(mobile: str, message: str, company_id: Optional[str] = None):
    """Send SMS via configured gateway"""
    try:
        if company_id:
            company_doc = db.companies.find_one({"id": company_id})
            if not company_doc or not company_doc.get("sms_enabled"):
                return False
            
            gateway = company_doc.get("sms_gateway", "textit")
            
            if gateway == "textit":
                username = company_doc.get("sms_username") or DEFAULT_SMS_USERNAME
                password = company_doc.get("sms_password") or DEFAULT_SMS_PASSWORD
                url = "https://www.textit.biz/sendmsg"
                params = {"id": username, "pw": password, "to": mobile, "text": message}
                response = requests.get(url, params=params, timeout=10)
                return response.status_code == 200
            
            elif gateway == "dialog":
                import hashlib
                username = company_doc.get("dialog_username")
                password = company_doc.get("dialog_password")
                mask = company_doc.get("dialog_mask")
                
                if not all([username, password, mask]):
                    logging.error("Dialog SMS: Missing credentials")
                    return False
                
                digest = hashlib.md5(password.encode()).hexdigest()
                url = "https://bulksms.dialog.lk/api/v2/send"
                payload = {
                    "user": username,
                    "digest": digest,
                    "mask": mask,
                    "destination": mobile,
                    "message": message
                }
                response = requests.post(url, json=payload, timeout=10)
                return response.status_code == 200
            
            elif gateway == "hutch":
                access_token = company_doc.get("hutch_access_token")
                if not access_token:
                    logging.error("Hutch SMS: Missing access token")
                    return False
                
                url = "https://bsms.hutch.lk/api/sms/send"
                headers = {"Authorization": f"Bearer {access_token}"}
                payload = {
                    "recipient": mobile,
                    "message": message
                }
                response = requests.post(url, json=payload, headers=headers, timeout=10)
                
                # If token expired, try to refresh
                if response.status_code == 401:
                    refresh_token = company_doc.get("hutch_refresh_token")
                    if refresh_token:
                        refresh_url = "https://bsms.hutch.lk/api/token/accessToken"
                        refresh_headers = {"Authorization": f"Bearer {refresh_token}"}
                        refresh_response = requests.post(refresh_url, headers=refresh_headers, timeout=10)
                        
                        if refresh_response.status_code == 200:
                            new_token = refresh_response.json().get("accessToken")
                            # Update token in database
                            db.companies.update_one(
                                {"id": company_id},
                                {"$set": {"hutch_access_token": new_token}}
                            )
                            # Retry send
                            headers["Authorization"] = f"Bearer {new_token}"
                            response = requests.post(url, json=payload, headers=headers, timeout=10)
                
                return response.status_code == 200
            
            elif gateway == "mobitel":
                app_id = company_doc.get("mobitel_app_id")
                app_key = company_doc.get("mobitel_app_key")
                client_id = company_doc.get("mobitel_client_id")
                
                if not all([app_id, app_key, client_id]):
                    logging.error("Mobitel SMS: Missing credentials")
                    return False
                
                url = "https://apphub.mobitel.lk/mobext/mapi/mspacesms/send"
                headers = {
                    "x-ibm-client-id": client_id,
                    "content-type": "application/json"
                }
                payload = {
                    "recipientMask": mobile,
                    "message": message,
                    "characterEncoding": "ascii",
                    "appID": app_id,
                    "appKey": app_key
                }
                response = requests.post(url, json=payload, headers=headers, timeout=10)
                return response.status_code == 200
            
        else:
            # System-wide gateway (for LOGIN OTP)
            url = "https://www.textit.biz/sendmsg"
            params = {"id": DEFAULT_SMS_USERNAME, "pw": DEFAULT_SMS_PASSWORD, "to": mobile, "text": message}
            response = requests.get(url, params=params, timeout=10)
            return response.status_code == 200
            
    except Exception as e:
        logging.error(f"SMS send error: {str(e)}")
        return False

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        # Check if this is an impersonation token
        is_impersonating = payload.get("is_impersonating", False)
        
        if is_impersonating:
            # For impersonation, we create a temporary user object with company context
            original_user = await db.users.find_one({"id": user_id}, {"_id": 0})
            if original_user is None or original_user.get("role") != "super_admin":
                raise HTTPException(status_code=401, detail="Invalid impersonation token")
            
            # Create a user object with company admin context
            impersonation_user = User(**{
                **original_user,
                "company_id": payload.get("company_id"),
                "role": "admin",  # Act as company admin
                "is_impersonating": True,
                "original_user_id": user_id,
                "can_edit_in_impersonation": payload.get("can_edit", False)
            })
            return impersonation_user
        else:
            user = await db.users.find_one({"id": user_id}, {"_id": 0})
            if user is None:
                raise HTTPException(status_code=401, detail="User not found")
            return User(**user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception as e:
        logging.error(f"Token validation error: {str(e)}")
        raise HTTPException(status_code=401, detail="Invalid token")

async def log_activity(company_id: str, user_id: str, user_name: str, action: str, details: str):
    log = ActivityLog(
        company_id=company_id,
        user_id=user_id,
        user_name=user_name,
        action=action,
        details=details
    )
    await db.activity_logs.insert_one(log.model_dump())

# ============= AUTH ENDPOINTS =============
@api_router.post("/auth/send-otp")
async def send_otp(request: OTPRequest):
    if len(request.mobile) != 10 or not request.mobile.isdigit():
        raise HTTPException(status_code=400, detail="Invalid mobile number")
    
    user = await db.users.find_one({"mobile": request.mobile}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    otp_code = str(random.randint(100000, 999999))
    
    otp_doc = {
        "mobile": request.mobile,
        "otp": otp_code,
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
        "verified": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.otps.insert_one(otp_doc)
    
    # Send SMS - LOGIN OTP always uses system-wide gateway (not company-specific)
    message = f"Your OTP for IT Signature ERP is: {otp_code}. Valid for 5 minutes."
    sms_sent = send_sms(request.mobile, message, None)  # None = use default system gateway
    
    # Log activity - OTP sent
    if user.get("company_id"):
        await log_activity(
            user["company_id"],
            user["id"],
            user["name"],
            "OTP_SENT",
            f"OTP sent to mobile {request.mobile}"
        )
    
    return {"message": "OTP sent successfully", "sms_sent": sms_sent}

@api_router.post("/auth/verify-otp")
async def verify_otp(request: OTPVerify):
    # If login_as is provided (role selection), allow already verified OTPs
    if request.login_as:
        otp_doc = await db.otps.find_one(
            {"mobile": request.mobile, "otp": request.otp},
            {"_id": 0},
            sort=[("created_at", -1)]
        )
    else:
        otp_doc = await db.otps.find_one(
            {"mobile": request.mobile, "otp": request.otp, "verified": False},
            {"_id": 0},
            sort=[("created_at", -1)]
        )
    
    # Check for multiple roles first to log invalid OTP with correct user context
    users = await db.users.find({"mobile": request.mobile}, {"_id": 0}).to_list(10)
    if not users:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Log invalid OTP attempt
    if not otp_doc:
        # Get first user for logging context
        first_user = users[0]
        if first_user.get("company_id"):
            await log_activity(
                first_user["company_id"],
                first_user["id"],
                first_user["name"],
                "INVALID_OTP",
                f"Invalid OTP attempt for mobile {request.mobile}"
            )
        raise HTTPException(status_code=400, detail="Invalid OTP")
    
    expires_at = datetime.fromisoformat(otp_doc["expires_at"])
    if datetime.now(timezone.utc) > expires_at:
        # Log expired OTP
        first_user = users[0]
        if first_user.get("company_id"):
            await log_activity(
                first_user["company_id"],
                first_user["id"],
                first_user["name"],
                "EXPIRED_OTP",
                f"Expired OTP attempt for mobile {request.mobile}"
            )
        raise HTTPException(status_code=400, detail="OTP expired")
    
    await db.otps.update_one(
        {"mobile": request.mobile, "otp": request.otp},
        {"$set": {"verified": True}}
    )
    
    has_super_admin = any(u["role"] == "super_admin" for u in users)
    has_company_role = any(u["role"] != "super_admin" for u in users)
    
    if has_super_admin and has_company_role and not request.login_as:
        return {
            "require_selection": True,
            "message": "This number has multiple access levels",
            "options": [{"value": "super_admin", "label": "Super Admin"}, {"value": "company", "label": "Company Portal"}]
        }
    
    # Select user
    if request.login_as == "super_admin":
        user = next((u for u in users if u["role"] == "super_admin"), None)
    elif request.login_as == "company" or not has_super_admin:
        user = next((u for u in users if u["role"] != "super_admin"), None)
    else:
        user = users[0]
    
    if not user:
        raise HTTPException(status_code=404, detail="User access not found")
    
    # Update last login
    if user.get("company_id"):
        await db.companies.update_one(
            {"id": user["company_id"]},
            {"$set": {"last_login": datetime.now(timezone.utc).isoformat()}}
        )
        
        # Log successful login
        await log_activity(
            user["company_id"],
            user["id"],
            user["name"],
            "LOGIN_SUCCESS",
            f"User logged in successfully"
        )
    
    token = create_access_token({"user_id": user["id"], "role": user["role"], "company_id": user.get("company_id")})
    
    return {"token": token, "user": user, "require_selection": False}

@api_router.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# ============= SUPER ADMIN ENDPOINTS =============
@api_router.post("/superadmin/companies", response_model=Company)
async def create_company(company: CompanyCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    existing = await db.companies.find_one({"admin_mobile": company.admin_mobile})
    if existing:
        raise HTTPException(status_code=400, detail="Mobile number already registered")
    
    company_obj = Company(**company.model_dump())
    await db.companies.insert_one(company_obj.model_dump())
    
    # Create admin user
    admin_user = User(
        company_id=company_obj.id,
        employee_id=f"ADMIN-{company_obj.id[:8]}",
        mobile=company.admin_mobile,
        name=company.admin_name,
        role="admin",
        join_date=datetime.now(timezone.utc).date().isoformat()
    )
    await db.users.insert_one(admin_user.model_dump())
    
    # Send SMS
    message = f"Welcome to IT Signature ERP! Your company '{company.name}' has been created. Login with mobile {company.admin_mobile}. URL: https://admin-sms-portal.preview.emergentagent.com"
    send_sms(company.admin_mobile, message)
    
    await log_activity("SUPER_ADMIN", current_user.id, current_user.name, "CREATE_COMPANY", f"Created company: {company.name}")
    
    return company_obj

@api_router.get("/superadmin/companies", response_model=List[Company])
async def get_all_companies(current_user: User = Depends(get_current_user)):
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    companies = await db.companies.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return companies

@api_router.get("/superadmin/companies/{company_id}")
async def get_company(company_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Get settings to include logo and favicon
    settings = await db.settings.find_one({"company_id": company_id}, {"_id": 0})
    
    company_data = Company(**company).model_dump()
    if settings:
        company_data["logo"] = settings.get("company_logo")
        company_data["favicon"] = settings.get("favicon")
    
    return company_data

@api_router.put("/superadmin/companies/{company_id}/status")
async def update_company_status(company_id: str, status: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    if status not in ["active", "suspended", "pending"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.companies.update_one(
        {"id": company_id},
        {"$set": {"status": status}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    await log_activity("SUPER_ADMIN", current_user.id, current_user.name, "UPDATE_STATUS", f"Changed {company['name']} status to {status}")
    
    return {"message": f"Company status updated to {status}"}

@api_router.put("/superadmin/companies/{company_id}/sms")
async def update_company_sms(company_id: str, sms_settings: SMSSettings, current_user: User = Depends(get_current_user)):
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    result = await db.companies.update_one(
        {"id": company_id},
        {"$set": sms_settings.model_dump()}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    await log_activity("SUPER_ADMIN", current_user.id, current_user.name, "UPDATE_SMS", f"Updated SMS settings for {company['name']}")
    
    return {"message": "SMS settings updated"}


@api_router.put("/superadmin/companies/{company_id}/short-code")
async def update_company_short_code(company_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """Update company short code for fingerprint attendance"""
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    short_code = data.get("short_code", "").strip()
    
    if not short_code:
        raise HTTPException(status_code=400, detail="Short code cannot be empty")
    
    if len(short_code) > 20:
        raise HTTPException(status_code=400, detail="Short code must be max 20 characters")
    
    # Check if short code already exists for another company
    existing = await db.companies.find_one({"short_code": short_code, "id": {"$ne": company_id}})
    if existing:
        raise HTTPException(status_code=400, detail="This short code is already in use by another company")
    
    result = await db.companies.update_one(
        {"id": company_id},
        {"$set": {"short_code": short_code}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    await log_activity("SUPER_ADMIN", current_user.id, current_user.name, "UPDATE_SHORT_CODE", f"Updated short code for {company['name']} to '{short_code}'")
    
    return {"message": "Short code updated successfully"}

@api_router.get("/superadmin/companies/{company_id}/admins")
async def get_company_admins(company_id: str, current_user: User = Depends(get_current_user)):
    """Get all admin users for a specific company"""
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    # Find all users with admin role for this company
    admins = await db.users.find(
        {"company_id": company_id, "role": "admin"},
        {"_id": 0, "id": 1, "name": 1, "mobile": 1, "employee_id": 1}
    ).to_list(100)
    
    return admins

@api_router.post("/superadmin/companies/{company_id}/resend-url")
async def resend_company_url(company_id: str, admin_id: str, current_user: User = Depends(get_current_user)):
    """Re-send company URL to a specific admin via SMS"""
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    # Get company details
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Get admin details
    admin = await db.users.find_one(
        {"id": admin_id, "company_id": company_id, "role": "admin"},
        {"_id": 0}
    )
    if not admin:
        raise HTTPException(status_code=404, detail="Admin not found")
    
    # Send SMS using default system gateway (same as OTP)
    message = f"Your company portal: {company['name']}. Login with mobile {admin['mobile']} at: https://admin-sms-portal.preview.emergentagent.com"
    sms_sent = send_sms(admin['mobile'], message, None)  # None = use default system gateway
    
    if not sms_sent:
        raise HTTPException(status_code=500, detail="Failed to send SMS")
    
    await log_activity(
        "SUPER_ADMIN", 
        current_user.id, 
        current_user.name, 
        "RESEND_URL", 
        f"Re-sent company URL for {company['name']} to admin {admin['name']} ({admin['mobile']})"
    )
    
    return {
        "message": "URL sent successfully",
        "admin_name": admin['name'],
        "admin_mobile": admin['mobile']
    }

@api_router.get("/superadmin/admins")
async def get_super_admins(current_user: User = Depends(get_current_user)):
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    admins = await db.users.find({"role": "super_admin"}, {"_id": 0}).to_list(100)
    return admins

@api_router.post("/superadmin/admins")
async def create_super_admin(admin_data: dict, current_user: User = Depends(get_current_user)):
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    # Check if mobile already exists
    existing = await db.users.find_one({"mobile": admin_data["mobile"]})
    if existing:
        raise HTTPException(status_code=400, detail="Mobile number already registered")
    
    # Create super admin
    new_admin = User(
        company_id=None,
        employee_id=admin_data.get("employee_id"),
        mobile=admin_data["mobile"],
        name=capitalize_name(admin_data["name"]),
        role="super_admin",
        join_date=datetime.now(timezone.utc).date().isoformat(),
        can_full_access_companies=admin_data.get("can_full_access_companies", False)
    )
    
    access_type = "Full Access" if admin_data.get("can_full_access_companies", False) else "Read-only"
    await db.users.insert_one(new_admin.model_dump())
    await log_activity("SUPER_ADMIN", current_user.id, current_user.name, "CREATE_SUPER_ADMIN", f"Created super admin: {capitalize_name(admin_data['name'])} with {access_type} permission for company views")
    
    return new_admin

@api_router.delete("/superadmin/admins/{admin_id}")
async def delete_super_admin(admin_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    # Check if trying to delete self
    if admin_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    # Check if last super admin
    super_admin_count = await db.users.count_documents({"role": "super_admin"})
    if super_admin_count <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete the last super admin")
    
    admin = await db.users.find_one({"id": admin_id, "role": "super_admin"})
    if not admin:
        raise HTTPException(status_code=404, detail="Super admin not found")
    
    await db.users.delete_one({"id": admin_id})
    await log_activity("SUPER_ADMIN", current_user.id, current_user.name, "DELETE_SUPER_ADMIN", f"Deleted super admin: {admin['name']}")
    
    return {"message": "Super admin deleted successfully"}

@api_router.get("/superadmin/dashboard/stats")
async def get_superadmin_stats(current_user: User = Depends(get_current_user)):
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    total_companies = await db.companies.count_documents({})
    active_companies = await db.companies.count_documents({"status": "active"})
    pending_companies = await db.companies.count_documents({"status": "pending"})
    total_employees = await db.users.count_documents({"role": {"$ne": "super_admin"}})
    
    # Get companies with stats
    companies = await db.companies.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    company_stats = []
    
    for company in companies:
        emp_count = await db.users.count_documents({"company_id": company["id"]})
        
        company_stats.append({
            "company_id": company["id"],
            "name": company["name"],
            "admin_name": company["admin_name"],
            "admin_mobile": company["admin_mobile"],
            "status": company["status"],
            "employee_count": emp_count,
            "last_login": company.get("last_login"),
            "sms_enabled": company.get("sms_enabled", False),
            "invoicing_enabled": company.get("invoicing_enabled", False),
            "location_tracking_enabled": company.get("location_tracking_enabled", False),
            "created_at": company["created_at"]
        })
    
    return {
        "total_companies": total_companies,
        "active_companies": active_companies,
        "pending_companies": pending_companies,
        "total_employees": total_employees,
        "company_stats": company_stats
    }

@api_router.post("/superadmin/impersonate/{company_id}")
async def impersonate_company(company_id: str, current_user: User = Depends(get_current_user)):
    """Super admin impersonates a company to view/manage their portal"""
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    # Verify company exists
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Get super admin's full access permission
    super_admin_user = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    can_edit = super_admin_user.get("can_full_access_companies", False)
    
    # Create impersonation token
    token_data = {
        "user_id": current_user.id,
        "company_id": company_id,
        "is_impersonating": True,
        "can_edit": can_edit
    }
    token = create_access_token(token_data)
    
    # Log the impersonation
    await log_activity(
        company_id, 
        current_user.id, 
        current_user.name, 
        "IMPERSONATE_START", 
        f"Super admin {current_user.name} started viewing company portal (Access: {'Full' if can_edit else 'Read-only'})"
    )
    
    return {
        "token": token,
        "company_name": company["name"],
        "company_id": company_id,
        "can_edit": can_edit,
        "message": f"Now viewing {company['name']} portal"
    }

@api_router.post("/superadmin/exit-impersonation")
async def exit_impersonation(current_user: User = Depends(get_current_user)):
    """Exit impersonation and return to super admin view"""
    if not hasattr(current_user, 'is_impersonating') or not current_user.is_impersonating:
        raise HTTPException(status_code=400, detail="Not currently impersonating")
    
    # Get original super admin user
    original_user_id = getattr(current_user, 'original_user_id', current_user.id)
    original_user = await db.users.find_one({"id": original_user_id}, {"_id": 0})
    
    if not original_user:
        raise HTTPException(status_code=404, detail="Original user not found")
    
    # Log the exit
    await log_activity(
        current_user.company_id,
        original_user_id,
        original_user["name"],
        "IMPERSONATE_END",
        f"Super admin {original_user['name']} exited company portal view"
    )
    
    # Create regular super admin token
    token_data = {"user_id": original_user_id}
    token = create_access_token(token_data)
    
    return {
        "token": token,
        "message": "Returned to super admin view"
    }

@api_router.put("/superadmin/admins/{admin_id}")
async def update_super_admin(admin_id: str, updates: dict, current_user: User = Depends(get_current_user)):
    """Update super admin details including full access permission"""
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    admin = await db.users.find_one({"id": admin_id, "role": "super_admin"})
    if not admin:
        raise HTTPException(status_code=404, detail="Super admin not found")
    
    # Only update allowed fields
    allowed_fields = ["name", "mobile", "can_full_access_companies"]
    update_data = {k: v for k, v in updates.items() if k in allowed_fields}
    
    if "name" in update_data:
        update_data["name"] = capitalize_name(update_data["name"])
    
    if update_data:
        await db.users.update_one(
            {"id": admin_id},
            {"$set": update_data}
        )
    
    # Log the update
    details = f"Updated super admin: {admin['name']}"
    if "can_full_access_companies" in update_data:
        access_type = "Full Access" if update_data["can_full_access_companies"] else "Read-only"
        details += f", Company View Permission: {access_type}"
    
    await log_activity("SUPER_ADMIN", current_user.id, current_user.name, "UPDATE_SUPER_ADMIN", details)
    
    return {"message": "Super admin updated successfully"}

# Settings Model
class CompanySettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    office_start_time: str = "09:00"
    office_end_time: str = "17:00"
    saturday_enabled: bool = True
    saturday_type: str = "full"
    saturday_start_time: str = "09:00"
    saturday_end_time: str = "14:00"
    working_days_per_month: int = 26
    holidays: List[dict] = []
    invoice_address: Optional[str] = None
    invoice_mobile: Optional[str] = None
    invoice_hotline: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account_name: Optional[str] = None
    bank_account_number: Optional[str] = None
    bank_branch: Optional[str] = None
    tin: Optional[str] = None
    place_of_supply: Optional[str] = None
    branch_code: Optional[str] = None
    company_logo: Optional[str] = None
    favicon: Optional[str] = None
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class SettingsUpdate(BaseModel):
    office_start_time: Optional[str] = None
    office_end_time: Optional[str] = None
    saturday_enabled: Optional[bool] = None
    saturday_type: Optional[str] = None
    saturday_start_time: Optional[str] = None
    saturday_end_time: Optional[str] = None
    working_days_per_month: Optional[int] = None
    invoice_address: Optional[str] = None
    invoice_mobile: Optional[str] = None
    invoice_hotline: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account_name: Optional[str] = None
    bank_account_number: Optional[str] = None
    bank_branch: Optional[str] = None
    tin: Optional[str] = None
    place_of_supply: Optional[str] = None
    branch_code: Optional[str] = None

class Holiday(BaseModel):
    date: str
    name: str
    type: str = "public"

# ============= COMPANY ENDPOINTS =============
@api_router.get("/company/info")
async def get_company_info(current_user: User = Depends(get_current_user)):
    if current_user.role == "super_admin":
        raise HTTPException(status_code=400, detail="Not applicable for super admin")

    company = await db.companies.find_one({"id": current_user.company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    # Get settings to include logo, favicon, and VAT-related fields
    settings = await db.settings.find_one({"company_id": current_user.company_id}, {"_id": 0})

    company_data = Company(**company).model_dump()
    if settings:
        company_data["logo"] = settings.get("company_logo")
        company_data["favicon"] = settings.get("favicon")
        # Include VAT-related fields from settings
        company_data["tin"] = settings.get("tin")
        company_data["place_of_supply"] = settings.get("place_of_supply")
        company_data["branch_code"] = settings.get("branch_code")

    return company_data

@api_router.put("/company/info")
async def update_company_info(info: CompanyInfoUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    result = await db.companies.update_one(
        {"id": current_user.company_id},
        {"$set": {
            "name": info.name,
            "address": info.address,
            "contact_number": info.contact_number,
            "email": info.email,
            "company_info_completed": True
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    
    changes = f"Name: {info.name}, Address: {info.address}, Contact: {info.contact_number}, Email: {info.email}"
    await log_activity(current_user.company_id, current_user.id, current_user.name, "UPDATE_INFO", f"Updated company information - {changes}")
    
    return {"message": "Company information updated successfully"}

@api_router.get("/company/logs")
async def get_company_logs(limit: int = 100, current_user: User = Depends(get_current_user)):
    if current_user.role == "super_admin":
        raise HTTPException(status_code=400, detail="Not applicable for super admin")
    
    logs = await db.activity_logs.find(
        {"company_id": current_user.company_id},
        {"_id": 0}
    ).sort("timestamp", -1).limit(limit).to_list(limit)
    
    return logs

@api_router.get("/activity-logs")
async def get_activity_logs(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    action_type: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    if current_user.role == "super_admin":
        raise HTTPException(status_code=403, detail="Super admin cannot access company logs directly. View via company portal.")
    
    query = {"company_id": current_user.company_id}
    
    # Filter by date range
    if from_date and to_date:
        query["timestamp"] = {"$gte": from_date, "$lte": to_date}
    elif from_date:
        query["timestamp"] = {"$gte": from_date}
    elif to_date:
        query["timestamp"] = {"$lte": to_date}
    
    # Filter by action type
    if action_type:
        query["action"] = {"$regex": action_type, "$options": "i"}
    
    # Search in user_name, action, or details
    if search:
        query["$or"] = [
            {"user_name": {"$regex": search, "$options": "i"}},
            {"action": {"$regex": search, "$options": "i"}},
            {"details": {"$regex": search, "$options": "i"}}
        ]
    
    logs = await db.activity_logs.find(query, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    
    return logs

# ============= DASHBOARD ENDPOINTS =============
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    if current_user.role == "super_admin":
        raise HTTPException(status_code=400, detail="Not applicable for super admin")
    
    if current_user.role in ["admin", "manager", "accountant"]:
        # Admin/Manager/Accountant stats
        # Get all active employees (status=1 or not set, is_active=True) - without date filter yet
        today_str = datetime.now(timezone.utc).date().isoformat()
        all_active_employees = await db.users.find({
            "company_id": current_user.company_id,
            "role": {"$ne": "super_admin"},  # Exclude only super_admin, include all company users
            "$or": [
                {"is_active": True},  # is_active = True
                {"is_active": {"$exists": False}}  # or is_active field doesn't exist (default active)
            ]
        }).to_list(length=None)
        
        # Filter out employees with status=0 (deleted) and check join_date
        active_employees_today = [
            emp for emp in all_active_employees
            if emp.get("status", 1) != 0  # Exclude status=0 (deleted)
            and (not emp.get("join_date") or emp.get("join_date", "") <= today_str)  # Check join_date
        ]
        
        total_employees = len(active_employees_today)
        active_employee_ids = [emp["id"] for emp in active_employees_today]
        
        # Count attendance only for active employees who have joined
        total_attendance_today = await db.attendance.count_documents({
            "company_id": current_user.company_id,
            "date": today_str,
            "employee_id": {"$in": active_employee_ids}
        })
        
        pending_leaves = await db.leaves.count_documents({"company_id": current_user.company_id, "status": "pending"})
        pending_advances = await db.advances.count_documents({"company_id": current_user.company_id, "status": "pending"})
        
        # Recent activities
        recent_leaves = await db.leaves.find({"company_id": current_user.company_id}, {"_id": 0}).sort("applied_date", -1).limit(5).to_list(5)
        recent_advances = await db.advances.find({"company_id": current_user.company_id}, {"_id": 0}).sort("request_date", -1).limit(5).to_list(5)
        
        # Current month salary summary
        current_month = datetime.now(timezone.utc).strftime("%B")
        current_year = datetime.now(timezone.utc).year
        
        monthly_payrolls = await db.payroll.find({
            "company_id": current_user.company_id,
            "month": current_month,
            "year": current_year
        }, {"_id": 0}).to_list(1000)
        
        # Calculate monthly stats
        total_expected_salary = sum(p.get("expected_salary", 0) for p in monthly_payrolls)
        total_calculated_salary = sum(p.get("calculated_salary", 0) for p in monthly_payrolls)
        total_net_salary = sum(p.get("net_salary", 0) for p in monthly_payrolls)
        
        # Current month attendance summary (last 7 days)
        # Count only active employees who have joined
        from datetime import timedelta
        today = datetime.now(timezone.utc).date()
        last_7_days = [(today - timedelta(days=i)).isoformat() for i in range(6, -1, -1)]
        
        attendance_summary = []
        for date in last_7_days:
            # Get employees who had joined by this specific date and are active (status != 0)
            employees_by_date = [
                emp for emp in all_active_employees 
                if emp.get("status", 1) != 0  # Exclude status=0 (deleted)
                and (not emp.get("join_date") or emp.get("join_date", "") <= date)  # Check join_date
            ]
            
            # Count ALL attendance for this date (for this company)
            # This shows actual attendance recorded regardless of employee status
            count = await db.attendance.count_documents({
                "company_id": current_user.company_id,
                "date": date
            })
            
            # Total employees is the count who had joined by this date
            attendance_summary.append({
                "date": date,
                "count": count,
                "total_employees": len(employees_by_date)
            })
        
        return {
            "total_employees": total_employees,
            "attendance_today": total_attendance_today,
            "pending_leaves": pending_leaves,
            "pending_advances": pending_advances,
            "recent_leaves": recent_leaves,
            "recent_advances": recent_advances,
            "monthly_salary_summary": {
                "month": current_month,
                "year": current_year,
                "total_expected": total_expected_salary,
                "total_calculated": total_calculated_salary,
                "total_net": total_net_salary,
                "employee_count": len(monthly_payrolls)
            },
            "attendance_summary": attendance_summary
        }
    else:
        # Employee/Staff stats
        my_attendance = await db.attendance.count_documents({"company_id": current_user.company_id, "employee_id": current_user.id})
        my_leaves = await db.leaves.find({"company_id": current_user.company_id, "employee_id": current_user.id}, {"_id": 0}).to_list(100)
        my_advances = await db.advances.find({"company_id": current_user.company_id, "employee_id": current_user.id}, {"_id": 0}).to_list(100)
        my_payroll = await db.payroll.find({"company_id": current_user.company_id, "employee_id": current_user.id}, {"_id": 0}).sort("generated_at", -1).limit(1).to_list(1)
        
        # Check today's attendance
        today = datetime.now(timezone.utc).date().isoformat()
        today_attendance = await db.attendance.find_one({"company_id": current_user.company_id, "employee_id": current_user.id, "date": today}, {"_id": 0})
        
        return {
            "total_attendance_days": my_attendance,
            "total_leaves": len(my_leaves),
            "approved_leaves": len([l for l in my_leaves if l["status"] == "approved"]),
            "total_advances": len(my_advances),
            "approved_advances": sum(a["amount"] for a in my_advances if a["status"] == "approved"),
            "latest_payroll": my_payroll[0] if my_payroll else None,
            "today_attendance": today_attendance,
            "my_leaves": sorted(my_leaves, key=lambda x: x.get("request_date", ""), reverse=True),
            "my_advances": sorted(my_advances, key=lambda x: x.get("request_date", ""), reverse=True)
        }

# ============= EMPLOYEE ENDPOINTS =============
@api_router.get("/employees")
async def get_employees(include_pending_increments: bool = False, include_deleted: bool = False, current_user: User = Depends(get_current_user)):
    if current_user.role == "super_admin":
        raise HTTPException(status_code=403, detail="Super admin cannot access company employees")
    
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    # Build query - exclude super_admin role
    query = {"company_id": current_user.company_id, "role": {"$ne": "super_admin"}}
    
    # By default, only show active employees (status != 0 AND is_active != False)
    if not include_deleted:
        query["$and"] = [
            {"$or": [{"status": {"$ne": 0}}, {"status": {"$exists": False}}]},
            {"$or": [{"is_active": {"$ne": False}}, {"is_active": {"$exists": False}}]}
        ]
    else:
        # If include_deleted=True, only show deleted (status = 0 OR is_active = false)
        query["$or"] = [
            {"status": 0},
            {"is_active": False}
        ]
    
    employees = await db.users.find(query, {"_id": 0}).to_list(1000)
    
    # Optionally include pending increments
    if include_pending_increments:
        pending_increments = await db.increments.find(
            {"company_id": current_user.company_id, "status": "pending"},
            {"_id": 0}
        ).to_list(length=None)
        
        # Create a map of employee_id to pending increment
        pending_map = {inc["employee_id"]: inc for inc in pending_increments}
        
        # Add pending increment to each employee
        for emp in employees:
            emp["pending_increment"] = pending_map.get(emp["id"])
    
    return employees

@api_router.post("/employees")
async def create_employee(employee: UserCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    # Check if employee already exists
    existing = await db.users.find_one({"mobile": employee.mobile, "company_id": current_user.company_id})
    if existing:
        raise HTTPException(status_code=400, detail="Employee with this mobile number already exists")
    
    # Get company settings for default times
    settings = await db.settings.find_one({"company_id": current_user.company_id})
    default_start_time = settings.get("office_start_time", "09:00") if settings else "09:00"
    default_finish_time = settings.get("office_end_time", "17:00") if settings else "17:00"
    
    # Create new employee
    new_employee = User(
        id=str(uuid.uuid4()),
        company_id=current_user.company_id,
        employee_id=employee.employee_id or f"EMP-{str(uuid.uuid4())[:8]}",
        mobile=employee.mobile,
        name=capitalize_name(employee.name),
        role=employee.role,
        department=employee.department or "",
        position=employee.position or "",
        basic_salary=employee.basic_salary or 0,
        allowances=employee.allowances or 0,
        join_date=employee.join_date,
        start_time=employee.start_time or default_start_time,
        finish_time=employee.finish_time or default_finish_time,
        fixed_salary=employee.fixed_salary or False,
        is_active=True,
        created_at=datetime.now(timezone.utc).isoformat()
    )
    
    await db.users.insert_one(new_employee.model_dump())
    await log_activity(current_user.company_id, current_user.id, current_user.name, "CREATE_EMPLOYEE", f"Created employee: {capitalize_name(employee.name)}, Role: {employee.role}, Mobile: {employee.mobile}, Department: {employee.department or 'N/A'}")
    
    return new_employee

@api_router.put("/employees/{employee_id}")
async def update_employee(employee_id: str, updates: dict, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    # Check if employee exists and belongs to the same company
    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Update employee
    await db.users.update_one(
        {"id": employee_id},
        {"$set": updates}
    )
    
    await log_activity(current_user.company_id, current_user.id, current_user.name, "UPDATE_EMPLOYEE", f"Updated employee: {employee['name']}. Changes: {', '.join([f'{k}={v}' for k, v in updates.items() if k not in ['_id', 'created_at']])}")
    
    return {"message": "Employee updated successfully"}

@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if employee exists and belongs to the same company
    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Soft delete - mark as inactive
    await db.users.update_one(
        {"id": employee_id},
        {"$set": {"is_active": False}}
    )
    
    await log_activity(current_user.company_id, current_user.id, current_user.name, "DELETE_EMPLOYEE", f"Deleted employee: {employee['name']}, ID: {employee.get('employee_id', 'N/A')}, Role: {employee.get('role', 'N/A')}")
    
    return {"message": "Employee deleted successfully"}


@api_router.patch("/employees/{employee_id}/reactivate")
async def reactivate_employee(employee_id: str, current_user: User = Depends(get_current_user)):
    """Reactivate a deleted/inactive employee (set status = 1)"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    # Check if employee exists and belongs to the same company
    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Reactivate employee - set status to 1 (active)
    await db.users.update_one(
        {"id": employee_id},
        {"$set": {"status": 1, "is_active": True}}
    )
    
    await log_activity(
        current_user.company_id, 
        current_user.id, 
        current_user.name, 
        "REACTIVATE_EMPLOYEE", 
        f"Reactivated employee: {employee['name']}, ID: {employee.get('employee_id', 'N/A')}, Role: {employee.get('role', 'N/A')}"
    )
    
    return {"message": "Employee reactivated successfully"}



# ============= BULK EMPLOYEE IMPORT ENDPOINTS =============
@api_router.post("/employees/parse-bulk")
async def parse_bulk_employees(data: dict, current_user: User = Depends(get_current_user)):
    """Use AI to parse pasted employee data and return structured format"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        from dotenv import load_dotenv
        load_dotenv()
        
        pasted_text = data.get("text", "")
        if not pasted_text.strip():
            raise HTTPException(status_code=400, detail="No text provided")
        
        # Initialize Gemini chat with faster model
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        chat = LlmChat(
            api_key=api_key,
            session_id=str(uuid.uuid4()),
            system_message="""You are an expert at parsing employee data from various formats. 
Extract employee information and return ONLY a valid JSON array. Each employee should be an object with these fields:
- name: Full name (string)
- email: Email address (string, can be null)
- mobile: Phone number (string, can be null)
- role: Job role/position/designation (string, can be null)
- position: Position/title (string, can be null)
- department: Department (string, can be null)
- join_date: Join date in YYYY-MM-DD format (string, can be null)

Rules:
1. Return ONLY a JSON array, no other text or explanation
2. Extract as much information as possible from the text
3. If a field is not found, use null
4. Standardize phone numbers to remove spaces and special characters
5. Convert dates to YYYY-MM-DD format
6. If role and position seem similar, use the value for both fields

Example output format:
[{"name":"John Doe","email":"john@example.com","mobile":"0771234567","role":"Manager","position":"Manager","department":"IT","join_date":"2023-01-15"}]"""
        ).with_model("gemini", "gemini-2.0-flash")
        
        # Create user message
        user_message = UserMessage(text=f"Parse this employee data:\n\n{pasted_text}")
        
        # Get AI response
        response = await chat.send_message(user_message)
        
        # Parse the JSON response
        import json
        # Clean the response - remove markdown code blocks if present
        cleaned_response = response.strip()
        if cleaned_response.startswith("```json"):
            cleaned_response = cleaned_response[7:]
        if cleaned_response.startswith("```"):
            cleaned_response = cleaned_response[3:]
        if cleaned_response.endswith("```"):
            cleaned_response = cleaned_response[:-3]
        cleaned_response = cleaned_response.strip()
        
        parsed_employees = json.loads(cleaned_response)
        
        # Validate it's a list
        if not isinstance(parsed_employees, list):
            raise ValueError("AI response is not a list")
        
        return {"employees": parsed_employees, "count": len(parsed_employees)}
        
    except Exception as e:
        logging.error(f"Error parsing bulk employees: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to parse employee data: {str(e)}")


@api_router.post("/employees/bulk-import")
async def bulk_import_employees(data: BulkEmployeeImportRequest, current_user: User = Depends(get_current_user)):
    """Import multiple employees after admin confirmation and editing"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    try:
        # Get company settings for default times
        settings = await db.settings.find_one({"company_id": current_user.company_id})
        default_start_time = settings.get("office_start_time", "09:00") if settings else "09:00"
        default_finish_time = settings.get("office_end_time", "17:00") if settings else "17:00"
        
        imported_count = 0
        errors = []
        
        for idx, emp_data in enumerate(data.employees):
            try:
                # Validate required fields
                if not emp_data.get("name"):
                    errors.append({"index": idx, "error": "Name is required"})
                    continue
                
                if not emp_data.get("mobile") and not emp_data.get("email"):
                    errors.append({"index": idx, "error": "Mobile or email is required"})
                    continue
                
                # Check for duplicate mobile
                if emp_data.get("mobile"):
                    existing = await db.users.find_one({
                        "mobile": emp_data["mobile"],
                        "company_id": current_user.company_id
                    })
                    if existing:
                        errors.append({"index": idx, "name": emp_data.get("name"), "error": f"Employee with mobile {emp_data['mobile']} already exists"})
                        continue
                
                # Create new employee
                new_employee = User(
                    id=str(uuid.uuid4()),
                    company_id=current_user.company_id,
                    employee_id=emp_data.get("employee_id") or f"EMP-{str(uuid.uuid4())[:8]}",
                    mobile=emp_data.get("mobile", ""),
                    name=capitalize_name(emp_data["name"]),
                    role=emp_data.get("role", "employee"),
                    department=emp_data.get("department", ""),
                    position=emp_data.get("position", ""),
                    basic_salary=float(emp_data.get("basic_salary", 0)),
                    allowances=float(emp_data.get("allowances", 0)),
                    join_date=emp_data.get("join_date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
                    start_time=emp_data.get("start_time") or default_start_time,
                    finish_time=emp_data.get("finish_time") or default_finish_time,
                    fixed_salary=emp_data.get("fixed_salary", False),
                    is_active=True,
                    created_at=datetime.now(timezone.utc).isoformat()
                )
                
                await db.users.insert_one(new_employee.model_dump())
                imported_count += 1
                
                # Log activity
                await log_activity(
                    current_user.company_id,
                    current_user.id,
                    current_user.name,
                    "BULK_IMPORT_EMPLOYEE",
                    f"Bulk imported employee: {capitalize_name(emp_data['name'])}, Role: {emp_data.get('role', 'employee')}"
                )
                
            except Exception as e:
                errors.append({"index": idx, "name": emp_data.get("name"), "error": str(e)})
        
        return {
            "message": f"Successfully imported {imported_count} employees",
            "imported_count": imported_count,
            "errors": errors
        }
        
    except Exception as e:
        logging.error(f"Error in bulk import: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Bulk import failed: {str(e)}")


# ============= INCREMENT ENDPOINTS =============
@api_router.post("/employees/{employee_id}/increments")
async def add_increment(employee_id: str, increment_data: dict, current_user: User = Depends(get_current_user)):
    """Add salary increment for an employee"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    # Get employee
    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    old_salary = employee.get("basic_salary", 0)
    new_salary = increment_data["new_salary"]
    increment_amount = new_salary - old_salary
    
    # Check if effective date is current or future
    from datetime import datetime
    effective_month = increment_data["effective_from"]  # Format: "YYYY-MM"
    current_month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    # Determine status and whether to apply immediately
    if effective_month <= current_month:
        status = "active"
        should_apply_now = True
    else:
        status = "pending"
        should_apply_now = False
    
    # Create increment record
    increment = Increment(
        company_id=current_user.company_id,
        employee_id=employee_id,
        employee_name=employee["name"],
        old_salary=old_salary,
        new_salary=new_salary,
        increment_amount=increment_amount,
        effective_from=increment_data["effective_from"],
        reason=increment_data.get("reason", ""),
        status=status,
        created_by=current_user.id,
        created_by_name=current_user.name
    )
    
    await db.increments.insert_one(increment.model_dump())
    
    # Only update employee's basic salary if effective date is now or past
    if should_apply_now:
        await db.users.update_one(
            {"id": employee_id},
            {"$set": {"basic_salary": new_salary}}
        )
        status_text = "Applied immediately"
    else:
        status_text = f"Scheduled for {effective_month}"
    
    # Log activity with detailed information
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "ADD_INCREMENT",
        f"Added salary increment for {employee['name']}: Rs. {old_salary:,.2f}  Rs. {new_salary:,.2f} (Increase: Rs. {increment_amount:,.2f}). Effective from: {increment_data['effective_from']}. Status: {status_text}. Reason: {increment_data.get('reason', 'N/A')}"
    )
    
    return {"message": "Increment added successfully", "increment": increment, "status": status}

@api_router.get("/employees/{employee_id}/increments")
async def get_employee_increments(employee_id: str, current_user: User = Depends(get_current_user)):
    """Get all increments for an employee"""
    # Verify employee belongs to current company
    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    increments = await db.increments.find(
        {"employee_id": employee_id, "company_id": current_user.company_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(length=None)
    
    return increments

@api_router.get("/employees/{employee_id}/pending-increment")
async def get_employee_pending_increment(employee_id: str, current_user: User = Depends(get_current_user)):
    """Get pending increment for an employee (if any)"""
    pending = await db.increments.find_one(
        {"employee_id": employee_id, "company_id": current_user.company_id, "status": "pending"},
        {"_id": 0}
    )
    
    return pending if pending else None


@api_router.get("/increments/pending")
async def get_all_pending_increments(current_user: User = Depends(get_current_user)):
    """Get all pending increments for the company"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    pending_increments = await db.increments.find(
        {"company_id": current_user.company_id, "status": "pending"},
        {"_id": 0}
    ).to_list(length=None)
    
    # Return as a map with employee_id as key
    result = {}
    for inc in pending_increments:
        result[inc["employee_id"]] = inc
    
    return result


@api_router.get("/increments")
async def get_all_increments(current_user: User = Depends(get_current_user)):
    """Get all increments for the company (admin view)"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    increments = await db.increments.find(
        {"company_id": current_user.company_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(length=None)
    
    return increments

@api_router.post("/increments/activate-pending")
async def activate_pending_increments(current_user: User = Depends(get_current_user)):
    """Activate pending increments whose effective date has arrived"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    current_month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    # Find all pending increments whose effective date is now or past
    pending_increments = await db.increments.find({
        "company_id": current_user.company_id,
        "status": "pending",
        "effective_from": {"$lte": current_month}
    }).to_list(length=None)
    
    activated_count = 0
    for increment in pending_increments:
        # Update employee salary
        await db.users.update_one(
            {"id": increment["employee_id"]},
            {"$set": {"basic_salary": increment["new_salary"]}}
        )
        
        # Mark increment as active
        await db.increments.update_one(
            {"id": increment["id"]},
            {"$set": {"status": "active"}}
        )
        
        # Log activation
        await log_activity(
            increment["company_id"],
            current_user.id,
            current_user.name,
            "ACTIVATE_INCREMENT",
            f"Activated salary increment for {increment['employee_name']}: Rs. {increment['old_salary']:,.2f}  Rs. {increment['new_salary']:,.2f}"
        )
        
        activated_count += 1
    
    return {
        "message": f"Activated {activated_count} pending increment(s)",
        "activated_count": activated_count
    }


# ============= LOAN ENDPOINTS =============
@api_router.post("/loans")
async def create_loan(loan_data: dict, current_user: User = Depends(get_current_user)):
    """Create a new loan for an employee"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    # Validate employee exists
    employee = await db.users.find_one({
        "id": loan_data["employee_id"],
        "company_id": current_user.company_id
    })
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    loan = {
        "id": str(uuid.uuid4()),
        "company_id": current_user.company_id,
        "employee_id": loan_data["employee_id"],
        "employee_name": employee["name"],
        "amount": loan_data["amount"],
        "monthly_deduction": loan_data["monthly_deduction"],
        "start_month": loan_data["start_month"],  # Format: YYYY-MM
        "remaining_amount": loan_data["amount"],
        "status": "active",
        "notes": loan_data.get("notes", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.name
    }
    
    await db.loans.insert_one(loan)
    
    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "CREATE_LOAN",
        f"Created loan for {employee['name']} - Rs. {loan_data['amount']}"
    )
    
    return loan

@api_router.get("/loans")
async def get_loans(employee_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get all loans or loans for a specific employee"""
    query = {"company_id": current_user.company_id}
    
    if employee_id:
        query["employee_id"] = employee_id
    
    loans = await db.loans.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=None)
    
    return loans

@api_router.put("/loans/{loan_id}/status")
async def update_loan_status(loan_id: str, status_data: dict, current_user: User = Depends(get_current_user)):
    """Update loan status (mark as completed/cancelled)"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    loan = await db.loans.find_one({
        "id": loan_id,
        "company_id": current_user.company_id
    })
    
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    await db.loans.update_one(
        {"id": loan_id, "company_id": current_user.company_id},
        {"$set": {"status": status_data["status"]}}
    )
    
    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "UPDATE_LOAN_STATUS",
        f"Updated loan status to {status_data['status']} for {loan['employee_name']}"
    )
    
    return {"message": "Loan status updated successfully"}

@api_router.delete("/loans/{loan_id}")
async def delete_loan(loan_id: str, current_user: User = Depends(get_current_user)):
    """Delete a loan"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    loan = await db.loans.find_one({
        "id": loan_id,
        "company_id": current_user.company_id
    })
    
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    await db.loans.delete_one({"id": loan_id, "company_id": current_user.company_id})
    
    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "DELETE_LOAN",
        f"Deleted loan for {loan['employee_name']}"
    )
    
    return {"message": "Loan deleted successfully"}


# ============= ADVANCES ENDPOINTS =============
@api_router.post("/advances")
async def create_advance(advance_data: dict, current_user: User = Depends(get_current_user)):
    """Create an advance for an employee (admin/manager/accountant only)"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")

    # Get employee details
    employee_id = advance_data.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="Employee ID is required")

    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    advance = Advance(
        company_id=current_user.company_id,
        employee_id=employee_id,
        employee_name=employee["name"],
        amount=advance_data["amount"],
        reason=advance_data.get("reason", ""),
        repayment_months=advance_data.get("repayment_months", 1),
        status="approved",  # Directly approved, no workflow
        approved_by=current_user.name
    )

    await db.advances.insert_one(advance.model_dump())

    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "CREATE_ADVANCE",
        f"Applied advance of Rs. {advance_data['amount']} for {employee['name']}"
    )

    return advance

@api_router.get("/advances")
async def get_advances(current_user: User = Depends(get_current_user)):
    """Get all advances (admin/manager see all, employees see only their own)"""
    if current_user.role in ["admin", "manager", "accountant"]:
        query = {"company_id": current_user.company_id}
    else:
        query = {"company_id": current_user.company_id, "employee_id": current_user.id}
    
    advances = await db.advances.find(query, {"_id": 0}).sort("request_date", -1).to_list(length=None)
    
    return advances

@api_router.put("/advances/{advance_id}")
async def update_advance(advance_id: str, advance_data: dict, current_user: User = Depends(get_current_user)):
    """Update advance details (admin/manager/accountant only)"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")

    advance = await db.advances.find_one({
        "id": advance_id,
        "company_id": current_user.company_id
    })

    if not advance:
        raise HTTPException(status_code=404, detail="Advance not found")

    # Get employee details if employee_id is provided
    employee_id = advance_data.get("employee_id", advance["employee_id"])
    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    update_data = {
        "employee_id": employee_id,
        "employee_name": employee["name"],
        "amount": advance_data.get("amount", advance["amount"]),
        "reason": advance_data.get("reason", advance.get("reason", "")),
        "repayment_months": advance_data.get("repayment_months", advance.get("repayment_months", 1))
    }

    await db.advances.update_one(
        {"id": advance_id, "company_id": current_user.company_id},
        {"$set": update_data}
    )

    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "UPDATE_ADVANCE",
        f"Updated advance for {employee['name']} (Rs. {advance_data.get('amount', advance['amount'])})"
    )

    return {"message": "Advance updated successfully"}

@api_router.delete("/advances/{advance_id}")
async def delete_advance(advance_id: str, current_user: User = Depends(get_current_user)):
    """Delete an advance (admin/manager/accountant only)"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")

    advance = await db.advances.find_one({
        "id": advance_id,
        "company_id": current_user.company_id
    })

    if not advance:
        raise HTTPException(status_code=404, detail="Advance not found")

    await db.advances.delete_one({"id": advance_id, "company_id": current_user.company_id})

    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "DELETE_ADVANCE",
        f"Deleted advance for {advance['employee_name']} (Rs. {advance['amount']})"
    )

    return {"message": "Advance deleted successfully"}


# ============= LEAVES ENDPOINTS =============
@api_router.post("/leaves")
async def create_leave(leave_data: dict, current_user: User = Depends(get_current_user)):
    """Create a leave for an employee (admin/manager/accountant only)"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")

    # Get employee details
    employee_id = leave_data.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="Employee ID is required")

    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    leave = Leave(
        company_id=current_user.company_id,
        employee_id=employee_id,
        employee_name=employee["name"],
        leave_type=leave_data["leave_type"],
        from_date=leave_data["from_date"],
        to_date=leave_data["to_date"],
        reason=leave_data.get("reason", ""),
        status="approved",  # Directly approved, no workflow
        approved_by=current_user.name
    )

    await db.leaves.insert_one(leave.model_dump())

    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "CREATE_LEAVE",
        f"Applied {leave_data['leave_type']} leave for {employee['name']} from {leave_data['from_date']} to {leave_data['to_date']}"
    )

    return leave

@api_router.get("/leaves")
async def get_leaves(current_user: User = Depends(get_current_user)):
    """Get all leaves (admin/manager see all, employees see only their own)"""
    if current_user.role in ["admin", "manager", "accountant"]:
        query = {"company_id": current_user.company_id}
    else:
        query = {"company_id": current_user.company_id, "employee_id": current_user.id}
    
    leaves = await db.leaves.find(query, {"_id": 0}).sort("applied_date", -1).to_list(length=None)
    
    return leaves

@api_router.put("/leaves/{leave_id}")
async def update_leave(leave_id: str, leave_data: dict, current_user: User = Depends(get_current_user)):
    """Update leave details (admin/manager/accountant only)"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")

    leave = await db.leaves.find_one({
        "id": leave_id,
        "company_id": current_user.company_id
    })

    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")

    # Get employee details if employee_id is provided
    employee_id = leave_data.get("employee_id", leave["employee_id"])
    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    update_data = {
        "employee_id": employee_id,
        "employee_name": employee["name"],
        "leave_type": leave_data.get("leave_type", leave["leave_type"]),
        "from_date": leave_data.get("from_date", leave["from_date"]),
        "to_date": leave_data.get("to_date", leave["to_date"]),
        "reason": leave_data.get("reason", leave.get("reason", ""))
    }

    await db.leaves.update_one(
        {"id": leave_id, "company_id": current_user.company_id},
        {"$set": update_data}
    )

    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "UPDATE_LEAVE",
        f"Updated leave for {employee['name']} ({leave_data.get('leave_type', leave['leave_type'])} from {leave_data.get('from_date', leave['from_date'])} to {leave_data.get('to_date', leave['to_date'])})"
    )

    return {"message": "Leave updated successfully"}

@api_router.delete("/leaves/{leave_id}")
async def delete_leave(leave_id: str, current_user: User = Depends(get_current_user)):
    """Delete a leave (admin/manager/accountant only)"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")

    leave = await db.leaves.find_one({
        "id": leave_id,
        "company_id": current_user.company_id
    })

    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")

    await db.leaves.delete_one({"id": leave_id, "company_id": current_user.company_id})

    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "DELETE_LEAVE",
        f"Deleted leave for {leave['employee_name']} ({leave['leave_type']} from {leave['from_date']} to {leave['to_date']})"
    )

    return {"message": "Leave deleted successfully"}


# ============= EXTRA PAYMENTS ENDPOINTS =============
@api_router.post("/extra-payments")
async def create_extra_payment(payment_data: dict, current_user: User = Depends(get_current_user)):
    """Create extra payment for an employee for a specific month"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    # Validate employee exists
    employee = await db.users.find_one({
        "id": payment_data["employee_id"],
        "company_id": current_user.company_id
    })
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    payment = {
        "id": str(uuid.uuid4()),
        "company_id": current_user.company_id,
        "employee_id": payment_data["employee_id"],
        "employee_name": employee["name"],
        "month": payment_data["month"],  # Format: YYYY-MM
        "amount": payment_data["amount"],
        "reason": payment_data.get("reason", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.name
    }
    
    await db.extra_payments.insert_one(payment)
    
    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "CREATE_EXTRA_PAYMENT",
        f"Added extra payment for {employee['name']} - Rs. {payment_data['amount']} for {payment_data['month']}"
    )
    
    return payment

@api_router.get("/extra-payments")
async def get_extra_payments(employee_id: Optional[str] = None, month: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get extra payments"""
    query = {"company_id": current_user.company_id}
    
    if employee_id:
        query["employee_id"] = employee_id
    
    if month:
        query["month"] = month
    
    payments = await db.extra_payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=None)
    
    return payments

@api_router.put("/extra-payments/{payment_id}")
async def update_extra_payment(payment_id: str, payment_data: dict, current_user: User = Depends(get_current_user)):
    """Update extra payment"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    payment = await db.extra_payments.find_one({
        "id": payment_id,
        "company_id": current_user.company_id
    })
    
    if not payment:
        raise HTTPException(status_code=404, detail="Extra payment not found")
    
    update_data = {
        "amount": payment_data.get("amount", payment["amount"]),
        "reason": payment_data.get("reason", payment["reason"])
    }
    
    await db.extra_payments.update_one(
        {"id": payment_id, "company_id": current_user.company_id},
        {"$set": update_data}
    )
    
    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "UPDATE_EXTRA_PAYMENT",
        f"Updated extra payment for {payment['employee_name']}"
    )
    
    return {"message": "Extra payment updated successfully"}

@api_router.delete("/extra-payments/{payment_id}")
async def delete_extra_payment(payment_id: str, current_user: User = Depends(get_current_user)):
    """Delete extra payment"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    payment = await db.extra_payments.find_one({
        "id": payment_id,
        "company_id": current_user.company_id
    })
    
    if not payment:
        raise HTTPException(status_code=404, detail="Extra payment not found")
    
    await db.extra_payments.delete_one({"id": payment_id, "company_id": current_user.company_id})
    
    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "DELETE_EXTRA_PAYMENT",
        f"Deleted extra payment for {payment['employee_name']}"
    )
    
    return {"message": "Extra payment deleted successfully"}

# ============= INVOICING ENDPOINTS =============

# Customer endpoints
@api_router.post("/customers")
async def create_customer(customer_data: dict, current_user: User = Depends(get_current_user)):
    """Create a new customer"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    customer = Customer(
        company_id=current_user.company_id,
        name=customer_data["name"],
        company_name=customer_data.get("company_name"),
        email=customer_data.get("email"),
        phone=customer_data.get("phone"),
        whatsapp=customer_data.get("whatsapp"),
        city=customer_data.get("city"),
        address=customer_data.get("address"),
        tin=customer_data.get("tin"),
        bank_name=customer_data.get("bank_name"),
        bank_branch=customer_data.get("bank_branch"),
        bank_account_number=customer_data.get("bank_account_number"),
        bank_account_holder_name=customer_data.get("bank_account_holder_name")
    )
    
    await db.customers.insert_one(customer.model_dump())
    await log_activity(current_user.company_id, current_user.id, current_user.name, "CREATE_CUSTOMER", f"Created customer: {customer.name}")
    
    return customer.model_dump()

@api_router.get("/customers")
async def get_customers(include_deleted: bool = False, current_user: User = Depends(get_current_user)):
    """Get all customers for company"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"company_id": current_user.company_id}
    if not include_deleted:
        query["deleted"] = {"$ne": True}
    
    customers = await db.customers.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=None)
    return customers

@api_router.put("/customers/{customer_id}")
async def update_customer(customer_id: str, customer_data: dict, current_user: User = Depends(get_current_user)):
    """Update customer"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    customer = await db.customers.find_one({"id": customer_id, "company_id": current_user.company_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await db.customers.update_one(
        {"id": customer_id, "company_id": current_user.company_id},
        {"$set": customer_data}
    )
    
    await log_activity(current_user.company_id, current_user.id, current_user.name, "UPDATE_CUSTOMER", f"Updated customer: {customer_data.get('name', customer['name'])}")
    
    return {"message": "Customer updated successfully"}

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: User = Depends(get_current_user)):
    """Soft delete customer"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    customer = await db.customers.find_one({"id": customer_id, "company_id": current_user.company_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Soft delete
    await db.customers.update_one(
        {"id": customer_id, "company_id": current_user.company_id},
        {"$set": {
            "deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.name
        }}
    )
    await log_activity(current_user.company_id, current_user.id, current_user.name, "DELETE_CUSTOMER", f"Deleted customer: {customer['name']}")
    
    return {"message": "Customer deleted successfully"}

@api_router.put("/customers/{customer_id}/restore")
async def restore_customer(customer_id: str, current_user: User = Depends(get_current_user)):
    """Restore deleted customer"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    customer = await db.customers.find_one({"id": customer_id, "company_id": current_user.company_id, "deleted": True})
    if not customer:
        raise HTTPException(status_code=404, detail="Deleted customer not found")
    
    await db.customers.update_one(
        {"id": customer_id, "company_id": current_user.company_id},
        {"$set": {"deleted": False}, "$unset": {"deleted_at": "", "deleted_by": ""}}
    )
    await log_activity(current_user.company_id, current_user.id, current_user.name, "RESTORE_CUSTOMER", f"Restored customer: {customer['name']}")
    
    return {"message": "Customer restored successfully"}

# Product Category endpoints
@api_router.post("/product-categories")
async def create_category(category_data: dict, current_user: User = Depends(get_current_user)):
    """Create product category"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    category = ProductCategory(
        company_id=current_user.company_id,
        name=category_data["name"]
    )
    
    await db.product_categories.insert_one(category.model_dump())
    return category.model_dump()

@api_router.get("/product-categories")
async def get_categories(current_user: User = Depends(get_current_user)):
    """Get all product categories"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    categories = await db.product_categories.find({"company_id": current_user.company_id}, {"_id": 0}).sort("name", 1).to_list(length=None)
    return categories

# Product endpoints
@api_router.post("/products")
async def create_product(product_data: dict, current_user: User = Depends(get_current_user)):
    """Create a new product"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    product = Product(
        company_id=current_user.company_id,
        category_id=product_data.get("category_id"),
        name=product_data["name"],
        description=product_data.get("description"),
        price=product_data["price"],
        unit=product_data.get("unit", "pcs"),
        stock_quantity=product_data.get("stock_quantity", 0)
    )
    
    await db.products.insert_one(product.model_dump())
    await log_activity(current_user.company_id, current_user.id, current_user.name, "CREATE_PRODUCT", f"Created product: {product.name}")
    
    return product.model_dump()

@api_router.get("/products")
async def get_products(include_deleted: bool = False, current_user: User = Depends(get_current_user)):
    """Get all products for company"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"company_id": current_user.company_id}
    if not include_deleted:
        query["deleted"] = {"$ne": True}
    
    products = await db.products.find(query, {"_id": 0}).sort("name", 1).to_list(length=None)
    return products

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, product_data: dict, current_user: User = Depends(get_current_user)):
    """Update product"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    product = await db.products.find_one({"id": product_id, "company_id": current_user.company_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    await db.products.update_one(
        {"id": product_id, "company_id": current_user.company_id},
        {"$set": product_data}
    )
    
    await log_activity(current_user.company_id, current_user.id, current_user.name, "UPDATE_PRODUCT", f"Updated product: {product_data.get('name', product['name'])}")
    
    return {"message": "Product updated successfully"}

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: User = Depends(get_current_user)):
    """Soft delete product"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    product = await db.products.find_one({"id": product_id, "company_id": current_user.company_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Soft delete
    await db.products.update_one(
        {"id": product_id, "company_id": current_user.company_id},
        {"$set": {
            "deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.name
        }}
    )
    await log_activity(current_user.company_id, current_user.id, current_user.name, "DELETE_PRODUCT", f"Deleted product: {product['name']}")
    
    return {"message": "Product deleted successfully"}

@api_router.put("/products/{product_id}/restore")
async def restore_product(product_id: str, current_user: User = Depends(get_current_user)):
    """Restore deleted product"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    product = await db.products.find_one({"id": product_id, "company_id": current_user.company_id, "deleted": True})
    if not product:
        raise HTTPException(status_code=404, detail="Deleted product not found")
    
    await db.products.update_one(
        {"id": product_id, "company_id": current_user.company_id},
        {"$set": {"deleted": False}, "$unset": {"deleted_at": "", "deleted_by": ""}}
    )
    await log_activity(current_user.company_id, current_user.id, current_user.name, "RESTORE_PRODUCT", f"Restored product: {product['name']}")
    
    return {"message": "Product restored successfully"}

# Invoice endpoints
def generate_invoice_number(branch_code: str = "MAIN"):
    """Generate invoice number: YYMMM_QQQQ_XXXXX format as per Sri Lankan VAT requirements
    Example: 25JAN_MAIN_00001 (Year 2025, January, Main branch, Invoice #1)
    """
    now = datetime.now(timezone.utc)
    yy = now.strftime("%y")  # Last 2 digits of year
    mmm = now.strftime("%b").upper()  # First 3 letters of month in uppercase
    qqqq = branch_code.upper() if branch_code else "MAIN"  # Branch/unit code
    return f"{yy}{mmm}_{qqqq}"

@api_router.post("/invoices")
async def create_invoice(invoice_data: dict, current_user: User = Depends(get_current_user)):
    """Create a new invoice"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")

    # Get company info for branch code
    company = await db.companies.find_one({"id": current_user.company_id})
    branch_code = company.get("branch_code", "MAIN") if company else "MAIN"

    # Generate invoice number with sequence: YYMMM_QQQQ_XXXXX
    base_number = generate_invoice_number(branch_code)

    # Count invoices with same month prefix to get sequential number
    count = await db.invoices.count_documents({
        "company_id": current_user.company_id,
        "invoice_number": {"$regex": f"^{base_number}_"}
    })

    # Format: 25JAN_MAIN_00001
    invoice_number = f"{base_number}_{str(count + 1).zfill(5)}"

    # Process items
    items = []
    for item_data in invoice_data["items"]:
        # Convert quantity and unit_price to proper numeric types
        quantity = float(item_data["quantity"])
        unit_price = float(item_data["unit_price"])

        item = InvoiceItem(
            product_id=item_data.get("product_id"),
            product_name=item_data["product_name"],
            description=item_data.get("description"),
            quantity=quantity,
            unit_price=unit_price,
            total=quantity * unit_price
        )
        items.append(item.model_dump())

        # Reduce stock if product_id provided
        if item_data.get("product_id"):
            await db.products.update_one(
                {"id": item_data["product_id"], "company_id": current_user.company_id},
                {"$inc": {"stock_quantity": -quantity}}
            )

    # Calculate subtotal (net amount excluding VAT)
    subtotal = sum([item["total"] for item in items])

    # Calculate VAT (18%)
    vat_rate = 18.0
    vat_amount = round(subtotal * (vat_rate / 100))

    # Calculate total including VAT
    total = subtotal + vat_amount

    # Get place of supply from company or use provided value
    place_of_supply = invoice_data.get("place_of_supply") or company.get("place_of_supply", "")

    invoice = Invoice(
        company_id=current_user.company_id,
        customer_id=invoice_data["customer_id"],
        invoice_number=invoice_number,
        invoice_date=invoice_data.get("invoice_date", datetime.now(timezone.utc).date().isoformat()),
        due_date=invoice_data.get("due_date"),
        date_of_delivery=invoice_data.get("date_of_delivery"),
        place_of_supply=place_of_supply,
        items=items,
        subtotal=subtotal,
        vat_rate=vat_rate,
        vat_amount=vat_amount,
        total=total,
        total_in_words=invoice_data.get("total_in_words"),
        payment_mode=invoice_data.get("payment_mode"),
        notes=invoice_data.get("notes"),
        created_by=current_user.id,
        created_by_name=current_user.name
    )

    await db.invoices.insert_one(invoice.model_dump())
    await log_activity(current_user.company_id, current_user.id, current_user.name, "CREATE_INVOICE", f"Created invoice: {invoice_number}")

    return invoice.model_dump()

@api_router.put("/invoices/{invoice_id}")
async def update_invoice(invoice_id: str, invoice_data: dict, current_user: User = Depends(get_current_user)):
    """Update an existing invoice (for adding missing VAT fields like date_of_delivery, place_of_supply, payment_mode)"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")

    # Get existing invoice
    existing_invoice = await db.invoices.find_one({"id": invoice_id, "company_id": current_user.company_id})
    if not existing_invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    # Recalculate VAT if items are provided
    if "items" in invoice_data and invoice_data["items"]:
        items = []
        for item_data in invoice_data["items"]:
            quantity = float(item_data["quantity"])
            unit_price = float(item_data["unit_price"])

            item = InvoiceItem(
                product_id=item_data.get("product_id"),
                product_name=item_data["product_name"],
                description=item_data.get("description"),
                quantity=quantity,
                unit_price=unit_price,
                total=quantity * unit_price
            )
            items.append(item.model_dump())

        # Calculate subtotal (net amount excluding VAT)
        subtotal = sum([item["total"] for item in items])

        # Calculate VAT (18%)
        vat_rate = 18.0
        vat_amount = round(subtotal * (vat_rate / 100))

        # Calculate total including VAT
        total = subtotal + vat_amount

        invoice_data["items"] = items
        invoice_data["subtotal"] = subtotal
        invoice_data["vat_rate"] = vat_rate
        invoice_data["vat_amount"] = vat_amount
        invoice_data["total"] = total

    # Update the invoice
    await db.invoices.update_one(
        {"id": invoice_id, "company_id": current_user.company_id},
        {"$set": invoice_data}
    )

    await log_activity(current_user.company_id, current_user.id, current_user.name, "UPDATE_INVOICE", f"Updated invoice: {existing_invoice.get('invoice_number')}")

    return {"message": "Invoice updated successfully"}

@api_router.get("/invoices")
async def get_invoices(
    status: Optional[str] = None,
    customer_id: Optional[str] = None,
    include_deleted: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Get all invoices for company"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"company_id": current_user.company_id}
    if not include_deleted:
        query["deleted"] = {"$ne": True}
    if status:
        query["status"] = status
    if customer_id:
        query["customer_id"] = customer_id
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=None)
    return invoices

@api_router.get("/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Get single invoice details"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    invoice = await db.invoices.find_one({"id": invoice_id, "company_id": current_user.company_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Get customer details
    customer = await db.customers.find_one({"id": invoice["customer_id"]}, {"_id": 0})
    invoice["customer"] = customer
    
    # Get payments
    payments = await db.invoice_payments.find({"invoice_id": invoice_id}, {"_id": 0}).sort("payment_date", -1).to_list(length=None)
    invoice["payments"] = payments
    
    return invoice

@api_router.post("/invoices/{invoice_id}/payments")
async def add_payment(invoice_id: str, payment_data: dict, current_user: User = Depends(get_current_user)):
    """Add payment to invoice"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    invoice = await db.invoices.find_one({"id": invoice_id, "company_id": current_user.company_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    payment = Payment(
        invoice_id=invoice_id,
        amount=payment_data["amount"],
        payment_date=payment_data.get("payment_date", datetime.now(timezone.utc).date().isoformat()),
        payment_method=payment_data["payment_method"],
        notes=payment_data.get("notes"),
        created_by=current_user.id,
        created_by_name=current_user.name
    )
    
    await db.invoice_payments.insert_one(payment.model_dump())
    
    # Update invoice amount_paid and status
    new_amount_paid = invoice["amount_paid"] + payment_data["amount"]
    new_status = "paid" if new_amount_paid >= invoice["total"] else "partial"
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"amount_paid": new_amount_paid, "status": new_status}}
    )
    
    await log_activity(current_user.company_id, current_user.id, current_user.name, "ADD_PAYMENT", f"Added payment Rs {payment_data['amount']} to invoice {invoice['invoice_number']}")
    
    return {"message": "Payment added successfully"}

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Soft delete invoice"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    invoice = await db.invoices.find_one({"id": invoice_id, "company_id": current_user.company_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    await db.invoices.update_one(
        {"id": invoice_id, "company_id": current_user.company_id},
        {"$set": {
            "deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.name
        }}
    )
    await log_activity(current_user.company_id, current_user.id, current_user.name, "DELETE_INVOICE", f"Deleted invoice: {invoice['invoice_number']}")
    
    return {"message": "Invoice deleted successfully"}

@api_router.put("/invoices/{invoice_id}/restore")
async def restore_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Restore deleted invoice"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    invoice = await db.invoices.find_one({"id": invoice_id, "company_id": current_user.company_id, "deleted": True})
    if not invoice:
        raise HTTPException(status_code=404, detail="Deleted invoice not found")
    
    await db.invoices.update_one(
        {"id": invoice_id, "company_id": current_user.company_id},
        {"$set": {"deleted": False}, "$unset": {"deleted_at": "", "deleted_by": ""}}
    )
    await log_activity(current_user.company_id, current_user.id, current_user.name, "RESTORE_INVOICE", f"Restored invoice: {invoice['invoice_number']}")
    
    return {"message": "Invoice restored successfully"}

# Estimate endpoints
def generate_estimate_number():
    """Generate estimate number: EST-25-MMDD-XX"""
    now = datetime.now(timezone.utc)
    year = now.strftime("%y")
    month_day = now.strftime("%m%d")
    return f"EST-{year}-{month_day}"

@api_router.post("/estimates")
async def create_estimate(estimate_data: dict, current_user: User = Depends(get_current_user)):
    """Create a new estimate"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Generate estimate number with sequence
    base_number = generate_estimate_number()
    count = await db.estimates.count_documents({
        "company_id": current_user.company_id,
        "estimate_number": {"$regex": f"^{base_number}"}
    })
    estimate_number = f"{base_number}-{str(count + 1).zfill(2)}"
    
    # Process items
    items = []
    for item_data in estimate_data["items"]:
        item = InvoiceItem(
            product_id=item_data.get("product_id"),
            product_name=item_data["product_name"],
            description=item_data.get("description"),
            quantity=item_data["quantity"],
            unit_price=item_data["unit_price"],
            total=item_data["quantity"] * item_data["unit_price"],
            display_amounts=item_data.get("display_amounts", True)
        )
        items.append(item.model_dump())

    # Calculate subtotal only for items with display_amounts=True
    subtotal = sum([item["total"] for item in items if item.get("display_amounts", True)])
    
    estimate = Estimate(
        company_id=current_user.company_id,
        customer_id=estimate_data["customer_id"],
        estimate_number=estimate_number,
        estimate_date=estimate_data.get("estimate_date", datetime.now(timezone.utc).date().isoformat()),
        valid_until=estimate_data.get("valid_until"),
        items=items,
        subtotal=subtotal,
        total=subtotal,
        notes=estimate_data.get("notes"),
        display_total_amounts=estimate_data.get("display_total_amounts", True),
        created_by=current_user.id,
        created_by_name=current_user.name
    )
    
    await db.estimates.insert_one(estimate.model_dump())
    await log_activity(current_user.company_id, current_user.id, current_user.name, "CREATE_ESTIMATE", f"Created estimate: {estimate_number}")
    
    return estimate.model_dump()

@api_router.get("/estimates")
async def get_estimates(include_deleted: bool = False, current_user: User = Depends(get_current_user)):
    """Get all estimates for company"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"company_id": current_user.company_id}
    if not include_deleted:
        query["deleted"] = {"$ne": True}
    
    estimates = await db.estimates.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=None)
    return estimates

@api_router.get("/estimates/{estimate_id}")
async def get_estimate(estimate_id: str, current_user: User = Depends(get_current_user)):
    """Get single estimate details"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")

    estimate = await db.estimates.find_one({"id": estimate_id, "company_id": current_user.company_id}, {"_id": 0})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")

    # Get customer details
    customer = await db.customers.find_one({"id": estimate["customer_id"]}, {"_id": 0})
    estimate["customer"] = customer

    return estimate

@api_router.post("/estimates/{estimate_id}/convert")
async def convert_estimate_to_invoice(estimate_id: str, current_user: User = Depends(get_current_user)):
    """Convert estimate to invoice"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")

    estimate = await db.estimates.find_one({"id": estimate_id, "company_id": current_user.company_id})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")

    # Get company info for branch code and default place of supply
    company = await db.companies.find_one({"id": current_user.company_id})
    settings = await db.settings.find_one({"company_id": current_user.company_id})
    branch_code = settings.get("branch_code", "MAIN") if settings else "MAIN"
    place_of_supply = settings.get("place_of_supply", "") if settings else ""

    # Generate invoice number with new VAT format: YYMMM_QQQQ_XXXXX
    base_number = generate_invoice_number(branch_code)
    count = await db.invoices.count_documents({
        "company_id": current_user.company_id,
        "invoice_number": {"$regex": f"^{base_number}_"}
    })
    invoice_number = f"{base_number}_{str(count + 1).zfill(5)}"

    # Calculate subtotal from estimate items
    subtotal = sum([item["total"] for item in estimate["items"]])

    # Calculate VAT (18%)
    vat_rate = 18.0
    vat_amount = round(subtotal * (vat_rate / 100))

    # Calculate total including VAT
    total = subtotal + vat_amount

    # Create invoice from estimate with VAT calculations
    invoice = Invoice(
        company_id=estimate["company_id"],
        customer_id=estimate["customer_id"],
        invoice_number=invoice_number,
        invoice_date=datetime.now(timezone.utc).date().isoformat(),
        due_date=None,
        date_of_delivery=None,  # To be filled by user when editing
        place_of_supply=place_of_supply,
        items=estimate["items"],
        subtotal=subtotal,
        vat_rate=vat_rate,
        vat_amount=vat_amount,
        total=total,
        total_in_words=None,  # To be filled by user when editing
        payment_mode=None,  # To be filled by user when editing
        notes=estimate.get("notes"),
        created_by=current_user.id,
        created_by_name=current_user.name
    )

    await db.invoices.insert_one(invoice.model_dump())

    # Update estimate status
    await db.estimates.update_one(
        {"id": estimate_id},
        {"$set": {"status": "converted"}}
    )

    # Reduce stock for products
    for item in estimate["items"]:
        if item.get("product_id"):
            await db.products.update_one(
                {"id": item["product_id"], "company_id": current_user.company_id},
                {"$inc": {"stock_quantity": -item["quantity"]}}
            )

    await log_activity(current_user.company_id, current_user.id, current_user.name, "CONVERT_ESTIMATE", f"Converted estimate {estimate['estimate_number']} to invoice {invoice_number}")

    return invoice.model_dump()

@api_router.delete("/estimates/{estimate_id}")
async def delete_estimate(estimate_id: str, current_user: User = Depends(get_current_user)):
    """Soft delete estimate"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    estimate = await db.estimates.find_one({"id": estimate_id, "company_id": current_user.company_id})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")
    
    await db.estimates.update_one(
        {"id": estimate_id, "company_id": current_user.company_id},
        {"$set": {
            "deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.name
        }}
    )
    await log_activity(current_user.company_id, current_user.id, current_user.name, "DELETE_ESTIMATE", f"Deleted estimate: {estimate['estimate_number']}")

    return {"message": "Estimate deleted successfully"}

@api_router.put("/estimates/{estimate_id}")
async def update_estimate(estimate_id: str, estimate_data: dict, current_user: User = Depends(get_current_user)):
    """Update estimate"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")

    estimate = await db.estimates.find_one({"id": estimate_id, "company_id": current_user.company_id})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")

    # Process items to calculate totals
    items = []
    for item_data in estimate_data.get("items", []):
        quantity = float(item_data["quantity"])
        unit_price = float(item_data["unit_price"])
        item = {
            "product_id": item_data.get("product_id"),
            "product_name": item_data["product_name"],
            "description": item_data.get("description"),
            "quantity": quantity,
            "unit_price": unit_price,
            "total": quantity * unit_price,
            "display_amounts": item_data.get("display_amounts", True)
        }
        items.append(item)

    # Calculate subtotal only for items with display_amounts=True
    subtotal = sum([item["total"] for item in items if item.get("display_amounts", True)])

    update_data = {
        "customer_id": estimate_data.get("customer_id", estimate["customer_id"]),
        "estimate_date": estimate_data.get("estimate_date", estimate["estimate_date"]),
        "valid_until": estimate_data.get("valid_until", estimate["valid_until"]),
        "items": items,
        "subtotal": subtotal,
        "total": subtotal,
        "notes": estimate_data.get("notes", estimate.get("notes")),
        "display_total_amounts": estimate_data.get("display_total_amounts", estimate.get("display_total_amounts", True))
    }

    await db.estimates.update_one(
        {"id": estimate_id, "company_id": current_user.company_id},
        {"$set": update_data}
    )

    await log_activity(current_user.company_id, current_user.id, current_user.name, "UPDATE_ESTIMATE", f"Updated estimate: {estimate['estimate_number']}")

    return {"message": "Estimate updated successfully"}

@api_router.put("/estimates/{estimate_id}/restore")
async def restore_estimate(estimate_id: str, current_user: User = Depends(get_current_user)):
    """Restore deleted estimate"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")

    estimate = await db.estimates.find_one({"id": estimate_id, "company_id": current_user.company_id, "deleted": True})
    if not estimate:
        raise HTTPException(status_code=404, detail="Deleted estimate not found")

    await db.estimates.update_one(
        {"id": estimate_id, "company_id": current_user.company_id},
        {"$set": {"deleted": False}, "$unset": {"deleted_at": "", "deleted_by": ""}}
    )
    await log_activity(current_user.company_id, current_user.id, current_user.name, "RESTORE_ESTIMATE", f"Restored estimate: {estimate['estimate_number']}")

    return {"message": "Estimate restored successfully"}

@api_router.put("/estimates/{estimate_id}/approve")
async def approve_estimate(estimate_id: str, current_user: User = Depends(get_current_user)):
    """Approve an estimate (admin/manager/accountant only)"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")

    estimate = await db.estimates.find_one({"id": estimate_id, "company_id": current_user.company_id, "deleted": False})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")

    await db.estimates.update_one(
        {"id": estimate_id, "company_id": current_user.company_id},
        {"$set": {
            "approval_status": "approved",
            "approved_by": current_user.id,
            "approved_by_name": current_user.name,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }, "$unset": {
            "rejected_by": "",
            "rejected_by_name": "",
            "rejected_at": ""
        }}
    )
    await log_activity(current_user.company_id, current_user.id, current_user.name, "APPROVE_ESTIMATE", f"Approved estimate: {estimate['estimate_number']}")

    return {"message": "Estimate approved successfully"}

@api_router.put("/estimates/{estimate_id}/reject")
async def reject_estimate(estimate_id: str, current_user: User = Depends(get_current_user)):
    """Reject an estimate (admin/manager/accountant only)"""
    if current_user.role not in ["admin", "manager", "accountant", "employee", "staff_member"]:
        raise HTTPException(status_code=403, detail="Access denied")

    estimate = await db.estimates.find_one({"id": estimate_id, "company_id": current_user.company_id, "deleted": False})
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")

    await db.estimates.update_one(
        {"id": estimate_id, "company_id": current_user.company_id},
        {"$set": {
            "approval_status": "rejected",
            "rejected_by": current_user.id,
            "rejected_by_name": current_user.name,
            "rejected_at": datetime.now(timezone.utc).isoformat()
        }, "$unset": {
            "approved_by": "",
            "approved_by_name": "",
            "approved_at": ""
        }}
    )
    await log_activity(current_user.company_id, current_user.id, current_user.name, "REJECT_ESTIMATE", f"Rejected estimate: {estimate['estimate_number']}")

    return {"message": "Estimate rejected successfully"}

# Company invoice settings
@api_router.put("/company/invoice-settings")
async def update_invoice_settings(settings_data: dict, current_user: User = Depends(get_current_user)):
    """Update company invoice settings"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    await db.companies.update_one(
        {"id": current_user.company_id},
        {"$set": {
            "invoice_address": settings_data.get("address"),
            "invoice_mobile": settings_data.get("mobile"),
            "invoice_hotline": settings_data.get("hotline"),
            "bank_name": settings_data.get("bank_name"),
            "bank_account_name": settings_data.get("bank_account_name"),
            "bank_account_number": settings_data.get("bank_account_number")
        }}
    )
    
    await log_activity(current_user.company_id, current_user.id, current_user.name, "UPDATE_INVOICE_SETTINGS", "Updated company invoice settings")
    
    return {"message": "Invoice settings updated successfully"}

# Super admin - toggle invoicing
@api_router.put("/superadmin/companies/{company_id}/invoicing")
async def toggle_invoicing(company_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """Enable/disable invoicing for company"""
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    await db.companies.update_one(
        {"id": company_id},
        {"$set": {"invoicing_enabled": data["enabled"]}}
    )
    
    return {"message": f"Invoicing {'enabled' if data['enabled'] else 'disabled'} successfully"}


@api_router.put("/superadmin/companies/{company_id}/location-tracking")
async def toggle_location_tracking(company_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """Enable/disable location tracking for company"""
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    await db.companies.update_one(
        {"id": company_id},
        {"$set": {"location_tracking_enabled": data["enabled"]}}
    )
    
    await log_activity(
        company_id,
        current_user.id,
        current_user.name,
        "TOGGLE_LOCATION_TRACKING",
        f"Location tracking {'enabled' if data['enabled'] else 'disabled'} for company"
    )
    
    return {"message": f"Location tracking {'enabled' if data['enabled'] else 'disabled'} successfully"}


# ============= ATTENDANCE ENDPOINTS =============
@api_router.get("/attendance")
async def get_attendance(
    employee_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    limit: Optional[int] = 100,
    current_user: User = Depends(get_current_user)
):
    if current_user.role == "super_admin":
        raise HTTPException(status_code=403, detail="Super admin cannot access company attendance")
    
    query = {"company_id": current_user.company_id}
    
    # Filter by employee if provided
    if employee_id:
        query["employee_id"] = employee_id
    elif current_user.role not in ["admin", "manager", "accountant"]:
        # Regular employees can only see their own attendance
        query["employee_id"] = current_user.id
    
    # Filter by date range if provided
    if from_date and to_date:
        query["date"] = {"$gte": from_date, "$lte": to_date}
    elif from_date:
        query["date"] = {"$gte": from_date}
    elif to_date:
        query["date"] = {"$lte": to_date}
    else:
        # Default: show last 30 days only if no date filter provided
        from datetime import datetime, timedelta
        thirty_days_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        query["date"] = {"$gte": thirty_days_ago}
    
    # Limit results
    attendance = await db.attendance.find(query, {"_id": 0}).sort("date", -1).limit(min(limit, 200)).to_list(length=None)
    
    # Check if each record has edit history (add flag)
    attendance_ids = [att["id"] for att in attendance]
    history_counts = await db.attendance_history.aggregate([
        {"$match": {"attendance_id": {"$in": attendance_ids}, "company_id": current_user.company_id}},
        {"$group": {"_id": "$attendance_id", "count": {"$sum": 1}}}
    ]).to_list(length=None)
    
    history_map = {item["_id"]: item["count"] for item in history_counts}
    
    # Enrich with employee profile pictures
    employee_ids = list(set([att.get("employee_id") for att in attendance if att.get("employee_id")]))
    employees = await db.users.find({"id": {"$in": employee_ids}}, {"id": 1, "profile_pic": 1, "_id": 0}).to_list(length=None)
    employee_profile_map = {emp["id"]: emp.get("profile_pic") for emp in employees}
    
    for att in attendance:
        att["has_history"] = att["id"] in history_map
        att["history_count"] = history_map.get(att["id"], 0)
        # Add profile picture
        att["profile_pic"] = employee_profile_map.get(att.get("employee_id"))
    
    return attendance

@api_router.post("/attendance")
async def add_manual_attendance(attendance_data: dict, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    # Validate date is not in the future
    from datetime import date
    try:
        attendance_date = date.fromisoformat(attendance_data["date"])
        today = date.today()
        if attendance_date > today:
            raise HTTPException(status_code=400, detail="Cannot add attendance for future dates")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Validate employee belongs to same company
    employee = await db.users.find_one({"id": attendance_data["employee_id"], "company_id": current_user.company_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Check how many attendance records exist for this date (allow up to 10 per day)
    existing_count = await db.attendance.count_documents({
        "company_id": current_user.company_id,
        "employee_id": attendance_data["employee_id"],
        "date": attendance_data["date"]
    })
    
    if existing_count >= 10:
        raise HTTPException(
            status_code=400, 
            detail=f"Daily limit exceeded. Maximum 10 attendance records per day allowed. Current count: {existing_count}"
        )
    
    # Create attendance record
    # Combine date with times to create ISO datetime strings
    check_in_datetime = None
    check_out_datetime = None
    
    if attendance_data.get("check_in"):
        check_in_datetime = f"{attendance_data['date']}T{attendance_data['check_in']}:00"
    
    if attendance_data.get("check_out"):
        check_out_datetime = f"{attendance_data['date']}T{attendance_data['check_out']}:00"
    
    new_attendance = {
        "id": str(uuid.uuid4()),
        "company_id": current_user.company_id,
        "employee_id": attendance_data["employee_id"],
        "employee_name": capitalize_name(employee["name"]),
        "date": attendance_data["date"],
        "check_in": check_in_datetime,
        "check_out": check_out_datetime,
        "status": attendance_data.get("status", "present"),
        "leave_type": attendance_data.get("leave_type"),
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Store a copy for response before inserting (to avoid _id field)
    attendance_response = new_attendance.copy()
    
    await db.attendance.insert_one(new_attendance)
    await log_activity(current_user.company_id, current_user.id, current_user.name, "ADD_ATTENDANCE", f"Added attendance for {capitalize_name(employee['name'])} on {attendance_data['date']}, Status: {attendance_data.get('status', 'present')}, Check-in: {attendance_data.get('check_in', 'N/A')}, Check-out: {attendance_data.get('check_out', 'N/A')}")
    
    return {"message": "Attendance added successfully", "attendance": attendance_response}

@api_router.put("/attendance/{attendance_id}")
async def update_attendance(attendance_id: str, attendance_data: dict, current_user: User = Depends(get_current_user)):
    """Update attendance record (e.g., add check-out time)"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    # Get existing attendance
    attendance = await db.attendance.find_one({"id": attendance_id, "company_id": current_user.company_id})
    if not attendance:
        raise HTTPException(status_code=404, detail="Attendance not found")
    
    # Store old values for history
    old_check_in = attendance.get("check_in")
    old_check_out = attendance.get("check_out")
    
    # Update fields
    update_data = {}
    changes = []
    
    if "check_in" in attendance_data:
        new_check_in = f"{attendance['date']}T{attendance_data['check_in']}:00"
        update_data["check_in"] = new_check_in
        if old_check_in != new_check_in:
            changes.append(f"Check-in: {old_check_in or 'None'}  {new_check_in}")
    
    if "check_out" in attendance_data and attendance_data["check_out"]:
        new_check_out = f"{attendance['date']}T{attendance_data['check_out']}:00"
        update_data["check_out"] = new_check_out
        if old_check_out != new_check_out:
            changes.append(f"Check-out: {old_check_out or 'None'}  {new_check_out}")
    
    if "status" in attendance_data:
        update_data["status"] = attendance_data["status"]
    
    if "leave_type" in attendance_data:
        update_data["leave_type"] = attendance_data["leave_type"]
    
    if update_data:
        await db.attendance.update_one(
            {"id": attendance_id},
            {"$set": update_data}
        )
        
        # Save edit history
        if changes:
            history_record = {
                "id": str(uuid.uuid4()),
                "attendance_id": attendance_id,
                "company_id": current_user.company_id,
                "employee_id": attendance["employee_id"],
                "employee_name": attendance["employee_name"],
                "date": attendance["date"],
                "changes": ", ".join(changes),
                "old_check_in": old_check_in,
                "old_check_out": old_check_out,
                "new_check_in": update_data.get("check_in", old_check_in),
                "new_check_out": update_data.get("check_out", old_check_out),
                "edited_by": current_user.name,
                "edited_by_id": current_user.id,
                "edited_at": datetime.now(timezone.utc).isoformat()
            }
            await db.attendance_history.insert_one(history_record)
        
        # Log the update
        await log_activity(
            current_user.company_id,
            current_user.id,
            current_user.name,
            "UPDATE_ATTENDANCE",
            f"Updated attendance for {attendance['employee_name']} on {attendance['date']}: {', '.join(changes) if changes else 'No changes'}"
        )
    
    return {"message": "Attendance updated successfully"}

@api_router.get("/attendance/{attendance_id}/history")
async def get_attendance_history(attendance_id: str, current_user: User = Depends(get_current_user)):
    """Get edit history for an attendance record"""
    history = await db.attendance_history.find(
        {"attendance_id": attendance_id, "company_id": current_user.company_id},
        {"_id": 0}
    ).sort("edited_at", -1).to_list(length=None)
    
    return history


@api_router.put("/attendance/{attendance_id}/status")
async def update_attendance_status(attendance_id: str, status_data: dict, current_user: User = Depends(get_current_user)):
    """Update attendance status with history tracking"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    # Get existing attendance
    attendance = await db.attendance.find_one({
        "id": attendance_id,
        "company_id": current_user.company_id
    })
    
    if not attendance:
        raise HTTPException(status_code=404, detail="Attendance not found")
    
    old_status = attendance.get("status")
    new_status = status_data.get("status")
    
    if old_status == new_status:
        return {"message": "Status unchanged"}
    
    # Save to history
    history_entry = {
        "id": str(uuid.uuid4()),
        "attendance_id": attendance_id,
        "company_id": current_user.company_id,
        "employee_id": attendance["employee_id"],
        "employee_name": attendance["employee_name"],
        "field_changed": "status",
        "old_value": old_status,
        "new_value": new_status,
        "edited_by": current_user.name,
        "edited_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.attendance_history.insert_one(history_entry)
    
    # Update attendance
    await db.attendance.update_one(
        {"id": attendance_id},
        {"$set": {"status": new_status}}
    )
    
    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "UPDATE_ATTENDANCE_STATUS",
        f"Changed status for {attendance['employee_name']} on {attendance['date']} from {old_status} to {new_status}"
    )
    
    return {"message": "Status updated successfully"}

@api_router.delete("/attendance/{attendance_id}")
async def delete_attendance(attendance_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    # Check if attendance exists and belongs to same company
    attendance = await db.attendance.find_one({"id": attendance_id, "company_id": current_user.company_id})
    if not attendance:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    
    # Store in deleted_attendance collection
    deleted_record = {
        **attendance,
        "deleted_by": current_user.id,
        "deleted_by_name": current_user.name,
        "deleted_at": datetime.now(timezone.utc).isoformat()
    }
    await db.deleted_attendance.insert_one(deleted_record)
    
    # Delete from attendance
    await db.attendance.delete_one({"id": attendance_id})
    await log_activity(current_user.company_id, current_user.id, current_user.name, "DELETE_ATTENDANCE", f"Deleted attendance for {attendance.get('employee_name', 'employee')} on {attendance.get('date', 'N/A')}, Status: {attendance.get('status', 'N/A')}, Check-in: {attendance.get('check_in', 'N/A')}")
    
    return {"message": "Attendance deleted successfully"}

@api_router.get("/attendance/date/{date}")
async def get_attendance_by_date(date: str, current_user: User = Depends(get_current_user)):
    """Get all attendance records for a specific date"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    # Get all employees in company with salary info
    employees = await db.users.find(
        {"company_id": current_user.company_id, "role": {"$in": ["admin", "employee", "manager", "accountant", "staff_member"]}},
        {"_id": 0, "id": 1, "name": 1, "employee_id": 1, "profile_pic": 1, "basic_salary": 1, "allowances": 1, "fixed_salary": 1}
    ).to_list(length=None)
    
    # Get attendance for the date
    attendance_records = await db.attendance.find(
        {"company_id": current_user.company_id, "date": date},
        {"_id": 0}
    ).to_list(length=None)
    
    # Create maps
    attendance_map = {record["employee_id"]: record for record in attendance_records}
    employee_data_map = {emp["id"]: emp for emp in employees}
    
    # Get company settings for working hours and working days
    company = await db.companies.find_one({"id": current_user.company_id}, {"_id": 0})
    working_hours = company.get("working_hours", {})
    start_time = working_hours.get("start", "09:00")
    finish_time = working_hours.get("finish", "17:00")
    
    # Calculate working days dynamically for the month
    year_month = date[:7]  # Extract YYYY-MM from date
    year_int = int(year_month.split("-")[0])
    month_int = int(year_month.split("-")[1])
    
    # Get settings for holidays and Saturday configuration
    db_settings = await db.settings.find_one({"company_id": current_user.company_id})
    holidays = db_settings.get("holidays", []) if db_settings else []
    saturday_enabled = db_settings.get("saturday_enabled", True) if db_settings else True
    saturday_type = db_settings.get("saturday_type", "full") if db_settings else "full"
    
    # Calculate working days for this month
    working_days_result = calculate_working_days(year_int, month_int, holidays, saturday_enabled, saturday_type)
    working_days = working_days_result["working_days"]
    
    # Calculate expected minutes per day
    from datetime import datetime, timezone
    start_dt = datetime.strptime(start_time, "%H:%M")
    finish_dt = datetime.strptime(finish_time, "%H:%M")
    expected_minutes_per_day = int((finish_dt - start_dt).total_seconds() / 60)
    
    # Build complete attendance list with earnings
    all_attendance = []
    total_earnings = 0
    
    for employee in employees:
        emp_data = employee_data_map[employee["id"]]
        basic_salary = emp_data.get("basic_salary", 0)
        allowances = emp_data.get("allowances", 0)
        fixed_salary = emp_data.get("fixed_salary", False)
        
        # Calculate salary per minute
        if fixed_salary:
            # For fixed salary: (basic_salary / 30 days / expected_minutes_per_day)
            salary_per_minute = basic_salary / 30 / expected_minutes_per_day if expected_minutes_per_day > 0 else 0
        else:
            # For non-fixed: basic_salary is already total per month, divide by total work minutes
            total_work_minutes_per_month = expected_minutes_per_day * working_days  # Use actual working days from settings
            salary_per_minute = basic_salary / total_work_minutes_per_month if total_work_minutes_per_month > 0 else 0
        
        if employee["id"] in attendance_map:
            # Add salary and earnings to existing attendance record
            record = attendance_map[employee["id"]]
            record["profile_pic"] = emp_data.get("profile_pic")
            record["employee_id_display"] = emp_data.get("employee_id")
            
            # Calculate earnings for this day
            earnings = 0
            # Use naive datetime (no timezone) to match stored times
            now = datetime.now()
            today_str = now.strftime("%Y-%m-%d")
            
            if record.get("check_in") and record.get("check_out"):
                # Completed attendance - calculate based on actual hours
                try:
                    checkin_dt = datetime.fromisoformat(record["check_in"])
                    checkout_dt = datetime.fromisoformat(record["check_out"])
                    # Work with naive datetimes - they're all in local time
                    duration = checkout_dt - checkin_dt
                    minutes_worked = int(duration.total_seconds() / 60)
                    earnings = minutes_worked * salary_per_minute
                except:
                    pass
            elif record.get("check_in") and not record.get("check_out") and date == today_str:
                # Ongoing attendance for today - calculate up to now
                try:
                    from datetime import timedelta
                    checkin_dt = datetime.fromisoformat(record["check_in"])
                    # Server is in UTC, but check-ins are in Sri Lanka time (UTC+5:30)
                    # Add 5.5 hours to server time to get Sri Lanka time
                    now_srilanka = now + timedelta(hours=5, minutes=30)
                    duration = now_srilanka - checkin_dt
                    minutes_worked = int(duration.total_seconds() / 60)
                    earnings = minutes_worked * salary_per_minute
                except Exception as e:
                    pass
            
            record["earnings"] = round(earnings, 2)
            record["salary_per_minute"] = round(salary_per_minute, 2)
            total_earnings += earnings
            all_attendance.append(record)
        else:
            # Employee has no record for this date - mark as absent
            all_attendance.append({
                "id": None,
                "employee_id": employee["id"],
                "employee_name": employee["name"],
                "employee_id_display": emp_data.get("employee_id"),
                "profile_pic": emp_data.get("profile_pic"),
                "company_id": current_user.company_id,
                "date": date,
                "status": "absent",
                "check_in": None,
                "check_out": None,
                "earnings": 0,
                "salary_per_minute": round(salary_per_minute, 2)
            })
    
    return {
        "attendance": all_attendance, 
        "date": date, 
        "total": len(all_attendance),
        "total_earnings": round(total_earnings, 2)
    }

@api_router.get("/attendance/deleted")
async def get_deleted_attendance(current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    deleted_records = await db.deleted_attendance.find(
        {"company_id": current_user.company_id},
        {"_id": 0}
    ).sort("deleted_at", -1).to_list(1000)
    
    return deleted_records

# ============= PAYROLL ENDPOINTS =============
@api_router.post("/payroll/generate")
async def generate_payroll(payroll_data: dict, current_user: User = Depends(get_current_user)):
    """Generate payroll for a specific month"""
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, Manager or Accountant access required")
    
    month = payroll_data["month"]  # Format: "YYYY-MM"
    year, month_num = month.split("-")
    
    # Get all employees for the company
    employees = await db.users.find({
        "company_id": current_user.company_id,
        "role": {"$in": ["employee", "staff_member", "manager"]}
    }).to_list(length=None)
    
    payroll_records = []
    
    for employee in employees:
        # Get effective salary for this month (considers increment history)
        effective_salary = await get_effective_salary(
            employee["id"], 
            current_user.company_id, 
            month
        )
        
        # Get attendance for this employee for this month
        attendance_records = await db.attendance.find({
            "employee_id": employee["id"],
            "company_id": current_user.company_id,
            "date": {"$regex": f"^{month}"}
        }).to_list(length=None)
        
        # Calculate attendance metrics
        present_days = len([a for a in attendance_records if a.get("status") == "present"])
        leave_days = len([a for a in attendance_records if a.get("status") == "leave"])
        half_days = len([a for a in attendance_records if a.get("status") == "half_day"])
        
        # Calculate total hours worked
        total_hours = 0
        for record in attendance_records:
            if record.get("check_in") and record.get("check_out"):
                try:
                    check_in = datetime.fromisoformat(record["check_in"])
                    check_out = datetime.fromisoformat(record["check_out"])
                    hours = (check_out - check_in).total_seconds() / 3600
                    total_hours += hours
                except:
                    pass
        
        # Get company settings for working hours
        settings = await db.settings.find_one({"company_id": current_user.company_id})
        expected_hours_per_day = 8  # default
        if settings:
            try:
                start = datetime.strptime(settings.get("start_time", "09:00"), "%H:%M")
                finish = datetime.strptime(settings.get("finish_time", "17:00"), "%H:%M")
                expected_hours_per_day = (finish - start).total_seconds() / 3600
            except:
                pass
        
        # Calculate late days (simplified - you can enhance this)
        late_days = 0
        
        # Calculate deductions
        deductions = employee.get("deductions", 0)
        
        # Calculate allowances
        allowances = employee.get("allowances", 0)
        
        # Calculate gross salary (using effective salary for this month)
        gross_salary = effective_salary
        
        # Calculate net salary
        net_salary = gross_salary + allowances - deductions
        
        # Create payroll record
        payroll_record = {
            "id": str(uuid.uuid4()),
            "company_id": current_user.company_id,
            "employee_id": employee["id"],
            "employee_name": employee["name"],
            "month": month,
            "basic_salary": effective_salary,  # This is the key - uses effective salary
            "allowances": allowances,
            "deductions": deductions,
            "gross_salary": gross_salary,
            "net_salary": net_salary,
            "present_days": present_days,
            "leave_days": leave_days,
            "half_days": half_days,
            "late_days": late_days,
            "total_hours": round(total_hours, 2),
            "generated_by": current_user.name,
            "generated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Save to database
        await db.payroll.insert_one(payroll_record)
        payroll_records.append(payroll_record)
    
    # Log activity
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "GENERATE_PAYROLL",
        f"Generated payroll for {month} - {len(payroll_records)} employee(s)"
    )
    
    return {
        "message": f"Payroll generated for {len(payroll_records)} employee(s)",
        "month": month,
        "employee_count": len(payroll_records)
    }

@api_router.get("/payroll")
async def get_payroll(month: Optional[str] = None, employee_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get payroll records"""
    query = {"company_id": current_user.company_id}
    
    if month:
        query["month"] = month
    
    if employee_id:
        query["employee_id"] = employee_id
    
    payroll = await db.payroll.find(query, {"_id": 0}).sort("generated_at", -1).to_list(length=None)
    
    return payroll

@api_router.get("/payroll/months")
async def get_payroll_months(current_user: User = Depends(get_current_user)):
    """Get all months with employee data and calculate totals"""
    # Get all employees for company
    employees = await db.users.find({
        "company_id": current_user.company_id,
        "role": {"$in": ["employee", "staff_member", "manager"]}
    }).to_list(length=None)
    
    if not employees:
        return []
    
    # Get distinct months from attendance records
    pipeline = [
        {"$match": {"company_id": current_user.company_id}},
        {"$group": {"_id": {"$substr": ["$date", 0, 7]}}},
        {"$sort": {"_id": -1}}
    ]
    
    months_data = await db.attendance.aggregate(pipeline).to_list(length=None)
    
    # Also check current month even if no attendance yet
    from datetime import datetime
    current_month = datetime.now(timezone.utc).strftime("%Y-%m")
    month_set = set([m["_id"] for m in months_data])
    month_set.add(current_month)
    
    result = []
    for month_str in sorted(month_set, reverse=True):
        # Get detailed payroll for this month to calculate totals
        detailed_response = await get_detailed_payroll(month_str, current_user)
        
        # Calculate totals from detailed records
        total_net = sum([emp.get("net_salary", 0) for emp in detailed_response.get("employees", [])])
        
        result.append({
            "month": month_str,
            "total_salary": round(total_net, 2),  # Now showing net salary
            "employee_count": len(employees)
        })
    
    return result

@api_router.get("/payroll/detailed/{month}")
async def get_detailed_payroll(month: str, current_user: User = Depends(get_current_user)):
    """Get detailed salary breakdown for all employees in a month"""
    # Get company and settings
    company = await db.companies.find_one({"id": current_user.company_id}, {"_id": 0})
    db_settings = await db.settings.find_one({"company_id": current_user.company_id})
    working_hours_per_day = 8
    start_time = "09:00"
    finish_time = "17:00"
    
    if db_settings:
        start_time = db_settings.get("start_time", "09:00")
        finish_time = db_settings.get("finish_time", "17:00")
        try:
            start_dt = datetime.strptime(start_time, "%H:%M")
            finish_dt = datetime.strptime(finish_time, "%H:%M")
            working_hours_per_day = (finish_dt - start_dt).total_seconds() / 3600
        except:
            pass
    
    # Calculate working days dynamically based on calendar and holidays
    year, month_num = month.split("-")
    year_int = int(year)
    month_int = int(month_num)
    
    # Get holidays and Saturday settings
    holidays = db_settings.get("holidays", []) if db_settings else []
    saturday_enabled = db_settings.get("saturday_enabled", True) if db_settings else True
    saturday_type = db_settings.get("saturday_type", "full") if db_settings else "full"
    
    # Calculate working days for this month
    working_days_result = calculate_working_days(year_int, month_int, holidays, saturday_enabled, saturday_type)
    working_days = working_days_result["working_days"]
    
    print(f"DEBUG DETAILED PAYROLL: Month={month}, Calculated Working Days={working_days}")
    
    # Get all employees (include admin to match live payroll endpoint)
    employees = await db.users.find({
        "company_id": current_user.company_id,
        "role": {"$in": ["admin", "employee", "staff_member", "manager", "accountant"]}
    }).to_list(length=None)
    
    print(f"DEBUG DETAILED PAYROLL: Found {len(employees)} employees for company {current_user.company_id}")
    
    detailed_records = []
    
    for employee in employees:
        # Get effective salary (considers increments)
        basic_salary = await get_effective_salary(employee["id"], current_user.company_id, month)
        
        # Get attendance for this month
        attendance_records = await db.attendance.find({
            "employee_id": employee["id"],
            "company_id": current_user.company_id,
            "date": {"$regex": f"^{month}"}
        }).to_list(length=None)
        
        # Calculate attendance metrics
        present_days = len([a for a in attendance_records if a.get("status") == "present"])
        leave_days = len([a for a in attendance_records if a.get("status") == "leave"])
        half_days = len([a for a in attendance_records if a.get("status") == "half_day"])
        allowed_leaves = len([a for a in attendance_records if a.get("status") == "allowed_leave"])
        allowed_half_days = len([a for a in attendance_records if a.get("status") == "allowed_half_day"])
        
        # Calculate total attendance minutes
        total_attendance_minutes = 0
        now = datetime.now()
        today_str = now.strftime("%Y-%m-%d")
        
        # For present days with check-in/out
        for record in attendance_records:
            record_date = record.get("date", "")
            
            # For completed days (check-in and check-out both present)
            if record.get("check_in") and record.get("check_out"):
                try:
                    checkin_dt = datetime.fromisoformat(record["check_in"])
                    checkout_dt = datetime.fromisoformat(record["check_out"])
                    # Work with naive datetimes - all in local time
                    duration = checkout_dt - checkin_dt
                    total_attendance_minutes += int(duration.total_seconds() / 60)
                except:
                    pass
            # For today's ongoing attendance (checked in but not out yet) - only if viewing current month
            elif record_date == today_str and record.get("check_in") and not record.get("check_out") and month == now.strftime("%Y-%m"):
                try:
                    checkin_dt = datetime.fromisoformat(record["check_in"])
                    # Use naive datetime - compare local to local
                    duration = now - checkin_dt
                    total_attendance_minutes += int(duration.total_seconds() / 60)
                except:
                    pass
        
        # For allowed leaves, add full day minutes (counts as worked)
        minutes_per_day = working_hours_per_day * 60
        total_attendance_minutes += (allowed_leaves * minutes_per_day)
        total_attendance_minutes += (allowed_half_days * minutes_per_day * 0.5)
        
        # Calculate late minutes and deduction
        late_minutes = 0
        late_deduction = 0
        
        if not employee.get("fixed_salary", False):  # Only if NOT fixed salary
            # Calculate expected check-in time
            expected_checkin = datetime.strptime(start_time, "%H:%M").time()
            
            for record in attendance_records:
                if record.get("check_in") and record.get("status") == "present":
                    try:
                        checkin_dt = datetime.fromisoformat(record["check_in"])
                        checkin_time = checkin_dt.time()
                        
                        # Compare times
                        expected_minutes = expected_checkin.hour * 60 + expected_checkin.minute
                        actual_minutes = checkin_time.hour * 60 + checkin_time.minute
                        
                        if actual_minutes > expected_minutes:
                            late_minutes += (actual_minutes - expected_minutes)
                    except:
                        pass
            
            # Calculate late deduction
            if late_minutes > 0 and working_days > 0:
                salary_per_day = basic_salary / working_days
                salary_per_hour = salary_per_day / working_hours_per_day
                salary_per_minute = salary_per_hour / 60
                late_deduction = late_minutes * salary_per_minute
        
        # Get advances for this month
        advances = await db.advances.find({
            "employee_id": employee["id"],
            "company_id": current_user.company_id,
            "status": "approved",
            "request_date": {"$regex": f"^{month}"}
        }).to_list(length=None)
        
        total_advances = sum([adv.get("amount", 0) for adv in advances])
        
        # Get other deductions
        other_deductions = employee.get("deductions", 0)
        
        # Get allowances
        allowances = employee.get("allowances", 0)
        
        # Get extra payments for this month
        extra_payments = await db.extra_payments.find({
            "employee_id": employee["id"],
            "company_id": current_user.company_id,
            "month": month
        }).to_list(length=None)
        
        total_extra_payment = sum([ep.get("amount", 0) for ep in extra_payments])
        
        # Get active loans for this employee
        active_loans = await db.loans.find({
            "employee_id": employee["id"],
            "company_id": current_user.company_id,
            "status": "active"
        }).to_list(length=None)
        
        # Calculate loan deduction for this month
        loan_deduction = 0
        for loan in active_loans:
            # Calculate how many months have passed since loan start
            loan_start_month = loan.get("start_month", month)
            if loan_start_month <= month:
                loan_deduction += loan.get("monthly_deduction", 0)
        
        # Calculate earnings and gross based on salary type
        salary_per_minute = (basic_salary / working_days / working_hours_per_day / 60) if working_days > 0 else 0
        
        # Check if this is the current month
        current_month = datetime.now(timezone.utc).strftime("%Y-%m")
        is_current_month = (month == current_month)
        
        if employee.get("fixed_salary", False):
            if is_current_month:
                # For current month: pro-rate based on time passed
                import calendar
                now = datetime.now()
                days_in_month = calendar.monthrange(int(year), int(month_num))[1]
                current_day = now.day
                hours_in_month = days_in_month * 24
                hours_passed = (current_day - 1) * 24 + now.hour + now.minute / 60 + now.second / 3600
                
                # Pro-rata calculation based on time passed
                earnings_basic = (basic_salary / hours_in_month) * hours_passed
                earnings_allowances_prorated = (allowances / hours_in_month) * hours_passed
                earnings = earnings_basic
                # Gross = Basic earnings + Extra payments (WITHOUT allowances)
                gross_salary = earnings + total_extra_payment
                allowances_to_add = earnings_allowances_prorated
            else:
                # For past completed months: full salary regardless of attendance
                earnings = basic_salary
                # Gross = Basic + Extra payments (WITHOUT allowances)
                gross_salary = earnings + total_extra_payment
                allowances_to_add = allowances
        else:
            # Non-fixed salary: based on actual attendance minutes
            earnings = total_attendance_minutes * salary_per_minute
            # Gross = Earnings + Extra payments (WITHOUT allowances)
            gross_salary = earnings + total_extra_payment
            allowances_to_add = allowances
        
        # Calculate net salary
        # Net = Gross + Allowances - Deductions (WITH allowances)
        total_deductions = late_deduction + total_advances + other_deductions + loan_deduction
        net_salary = gross_salary + allowances_to_add - total_deductions
        
        print(f"DEBUG DETAILED PAYROLL EMPLOYEE: {employee['name']} - gross={gross_salary:.2f}, earnings={earnings:.2f}, minutes={total_attendance_minutes}")
        
        detailed_records.append({
            "employee_id": employee["id"],
            "employee_name": employee["name"],
            "position": employee.get("position", "Staff"),
            "profile_picture": employee.get("profile_pic"),
            "basic_salary": round(basic_salary, 2),
            "allowances": round(allowances_to_add, 2),
            "earnings": round(earnings, 2),
            "working_days": working_days,
            "present_days": present_days,
            "leave_days": leave_days,
            "half_days": half_days,
            "allowed_leaves": allowed_leaves,
            "allowed_half_days": allowed_half_days,
            "total_attendance_minutes": total_attendance_minutes,
            "late_minutes": late_minutes,
            "late_deduction": round(late_deduction, 2),
            "advances": round(total_advances, 2),
            "other_deductions": round(other_deductions, 2),
            "loan_deduction": round(loan_deduction, 2),
            "extra_payment": round(total_extra_payment, 2),
            "gross_salary": round(gross_salary, 2),
            "total_deductions": round(total_deductions, 2),
            "net_salary": round(net_salary, 2),
            "fixed_salary": employee.get("fixed_salary", False),
            "salary_per_minute": round(salary_per_minute, 2)
        })
    
    total_gross_calc = sum([r["gross_salary"] for r in detailed_records])
    print(f"DEBUG DETAILED PAYROLL: Calculated total_gross={total_gross_calc} from {len(detailed_records)} records")
    
    return {
        "month": month,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "employees": detailed_records,
        "total_gross": total_gross_calc,
        "total_net": sum([r["net_salary"] for r in detailed_records]),
        "total_deductions": sum([r["total_deductions"] for r in detailed_records]),
        "total_allowances": sum([r["allowances"] for r in detailed_records])
    }


@api_router.get("/payroll/live-current-month")
async def get_live_current_month_payroll(current_user: User = Depends(get_current_user)):
    """Get real-time payroll calculation for current month up to this second"""
    now = datetime.now()
    current_month = now.strftime("%Y-%m")
    
    # Get company settings
    settings = await db.settings.find_one({"company_id": current_user.company_id})
    working_hours_per_day = 8
    start_time = "09:00"
    finish_time = "17:00"
    
    if settings:
        start_time = settings.get("start_time", "09:00")
        finish_time = settings.get("finish_time", "17:00")
        try:
            start_dt = datetime.strptime(start_time, "%H:%M")
            finish_dt = datetime.strptime(finish_time, "%H:%M")
            working_hours_per_day = (finish_dt - start_dt).total_seconds() / 3600
        except:
            pass
    
    # Calculate working days dynamically for current month
    year, month_num = current_month.split("-")
    year_int = int(year)
    month_int = int(month_num)
    
    # Get holidays and Saturday settings
    holidays = settings.get("holidays", []) if settings else []
    saturday_enabled = settings.get("saturday_enabled", True) if settings else True
    saturday_type = settings.get("saturday_type", "full") if settings else "full"
    
    # Calculate working days for this month
    working_days_result = calculate_working_days(year_int, month_int, holidays, saturday_enabled, saturday_type)
    working_days = working_days_result["working_days"]
    
    print(f"DEBUG WORKING DAYS: Month={current_month}, Calculated Working Days={working_days}")
    
    # Get all employees (only if employee role, show own data; if admin/manager, show all)
    if current_user.role == "employee":
        employees = await db.users.find({
            "id": current_user.id,
            "company_id": current_user.company_id
        }).to_list(length=None)
    else:
        employees = await db.users.find({
            "company_id": current_user.company_id,
            "role": {"$in": ["admin", "employee", "staff_member", "manager", "accountant"]}
        }).to_list(length=None)
    
    print(f"DEBUG LIVE PAYROLL: Found {len(employees)} employees for company {current_user.company_id}")
    
    detailed_records = []
    today_total_earnings = 0  # Track today's earnings across all employees
    
    for employee in employees:
        # Get effective salary
        basic_salary = await get_effective_salary(employee["id"], current_user.company_id, current_month)
        
        # Get attendance for current month
        attendance_records = await db.attendance.find({
            "employee_id": employee["id"],
            "company_id": current_user.company_id,
            "date": {"$regex": f"^{current_month}"}
        }).to_list(length=None)
        
        # Calculate attendance metrics
        present_days = len([a for a in attendance_records if a.get("status") == "present"])
        leave_days = len([a for a in attendance_records if a.get("status") == "leave"])
        half_days = len([a for a in attendance_records if a.get("status") == "half_day"])
        allowed_leaves = len([a for a in attendance_records if a.get("status") == "allowed_leave"])
        allowed_half_days = len([a for a in attendance_records if a.get("status") == "allowed_half_day"])
        
        # Calculate total attendance minutes UP TO NOW
        total_attendance_minutes = 0
        today_minutes = 0  # Track today's minutes for this employee
        today_str = now.strftime("%Y-%m-%d")
        
        for record in attendance_records:
            record_date = record.get("date", "")
            
            # For completed days (check-in and check-out both present)
            if record.get("check_in") and record.get("check_out"):
                try:
                    checkin_dt = datetime.fromisoformat(record["check_in"])
                    checkout_dt = datetime.fromisoformat(record["check_out"])
                    # Work with naive datetimes - all in local time
                    duration = checkout_dt - checkin_dt
                    minutes_worked = int(duration.total_seconds() / 60)
                    total_attendance_minutes += minutes_worked
                    # If this is today's completed record, track it
                    if record_date == today_str:
                        today_minutes += minutes_worked
                except Exception as e:
                    pass
            # For today's ongoing attendance (checked in but not out yet)
            elif record_date == today_str and record.get("check_in") and not record.get("check_out"):
                try:
                    from datetime import timedelta
                    checkin_dt = datetime.fromisoformat(record["check_in"])
                    
                    # For total_attendance_minutes: Use NO timezone adjustment (match detailed payroll)
                    duration_no_tz = now - checkin_dt
                    minutes_no_tz = int(duration_no_tz.total_seconds() / 60)
                    total_attendance_minutes += minutes_no_tz
                    
                    # For today_minutes (used in Today Salary): Use timezone adjustment
                    # Server is in UTC, check-ins are in Sri Lanka time (UTC+5:30)
                    now_srilanka = now + timedelta(hours=5, minutes=30)
                    duration_with_tz = now_srilanka - checkin_dt
                    minutes_with_tz = int(duration_with_tz.total_seconds() / 60)
                    today_minutes += minutes_with_tz  # Use adjusted time for today's earnings
                except Exception as e:
                    pass
        
        # Add allowed leaves (count as worked time)
        minutes_per_day = working_hours_per_day * 60
        total_attendance_minutes += (allowed_leaves * minutes_per_day)
        total_attendance_minutes += (allowed_half_days * minutes_per_day * 0.5)
        
        # Calculate late minutes and deduction
        late_minutes = 0
        late_deduction = 0
        
        if not employee.get("fixed_salary", False):
            expected_checkin = datetime.strptime(start_time, "%H:%M").time()
            
            for record in attendance_records:
                if record.get("check_in") and record.get("status") == "present":
                    try:
                        checkin_dt = datetime.fromisoformat(record["check_in"])
                        checkin_time = checkin_dt.time()
                        
                        expected_minutes = expected_checkin.hour * 60 + expected_checkin.minute
                        actual_minutes = checkin_time.hour * 60 + checkin_time.minute
                        
                        if actual_minutes > expected_minutes:
                            late_minutes += (actual_minutes - expected_minutes)
                    except:
                        pass
            
            if late_minutes > 0 and working_days > 0:
                salary_per_day = basic_salary / working_days
                salary_per_hour = salary_per_day / working_hours_per_day
                salary_per_minute = salary_per_hour / 60
                late_deduction = late_minutes * salary_per_minute
        
        # Get advances for current month
        advances = await db.advances.find({
            "employee_id": employee["id"],
            "company_id": current_user.company_id,
            "status": "approved",
            "request_date": {"$regex": f"^{current_month}"}
        }).to_list(length=None)
        
        total_advances = sum([adv.get("amount", 0) for adv in advances])
        
        # Get other deductions
        other_deductions = employee.get("deductions", 0)
        
        # Get allowances
        allowances = employee.get("allowances", 0)
        
        # Get extra payments
        extra_payments = await db.extra_payments.find({
            "employee_id": employee["id"],
            "company_id": current_user.company_id,
            "month": current_month
        }).to_list(length=None)
        
        total_extra_payment = sum([ep.get("amount", 0) for ep in extra_payments])
        
        # Get active loans
        active_loans = await db.loans.find({
            "employee_id": employee["id"],
            "company_id": current_user.company_id,
            "status": "active"
        }).to_list(length=None)
        
        loan_deduction = 0
        for loan in active_loans:
            loan_start_month = loan.get("start_month", current_month)
            if loan_start_month <= current_month:
                loan_deduction += loan.get("monthly_deduction", 0)
        
        # Calculate earnings based on salary type
        salary_per_minute = (basic_salary / working_days / working_hours_per_day / 60) if working_days > 0 else 0
        
        if employee.get("fixed_salary", False):
            # Fixed salary for CURRENT MONTH: pro-rate based on time passed
            import calendar
            days_in_month = calendar.monthrange(int(year), int(month_num))[1]
            current_day = now.day
            hours_in_month = days_in_month * 24
            hours_passed = (current_day - 1) * 24 + now.hour + now.minute / 60 + now.second / 3600
            
            # Pro-rata calculation based on time passed
            earnings_basic = (basic_salary / hours_in_month) * hours_passed
            earnings_allowances_prorated = (allowances / hours_in_month) * hours_passed
            earnings = earnings_basic
            # Gross = Basic earnings + Extra payments (WITHOUT allowances)
            gross_salary = earnings + total_extra_payment
            # For fixed salary, use prorated allowances
            allowances_to_add = earnings_allowances_prorated
        else:
            # Non-fixed: based on actual minutes worked so far
            earnings = total_attendance_minutes * salary_per_minute
            # Gross = Earnings + Extra payments (WITHOUT allowances)
            gross_salary = earnings + total_extra_payment
            # For non-fixed, use full allowances
            allowances_to_add = allowances
        
        # Calculate net salary
        # Net = Gross + Allowances - Deductions (WITH allowances)
        total_deductions = late_deduction + total_advances + other_deductions + loan_deduction
        net_salary = gross_salary + allowances_to_add - total_deductions
        
        print(f"DEBUG LIVE PAYROLL EMPLOYEE: {employee['name']} - gross={gross_salary:.2f}, earnings={earnings:.2f}, minutes={total_attendance_minutes}")
        
        # Calculate today's earnings for this employee (only for non-fixed)
        today_earnings = today_minutes * salary_per_minute if not employee.get("fixed_salary", False) else 0
        today_total_earnings += today_earnings
        
        detailed_records.append({
            "employee_id": employee["id"],
            "employee_name": employee["name"],
            "position": employee.get("position", "Staff"),
            "profile_picture": employee.get("profile_pic"),
            "basic_salary": round(basic_salary, 2),
            "allowances": round(allowances_to_add, 2),
            "earnings": round(earnings, 2),
            "working_days": working_days,
            "present_days": present_days,
            "leave_days": leave_days,
            "half_days": half_days,
            "allowed_leaves": allowed_leaves,
            "allowed_half_days": allowed_half_days,
            "total_attendance_minutes": total_attendance_minutes,
            "late_minutes": late_minutes,
            "late_deduction": round(late_deduction, 2),
            "advances": round(total_advances, 2),
            "other_deductions": round(other_deductions, 2),
            "loan_deduction": round(loan_deduction, 2),
            "extra_payment": round(total_extra_payment, 2),
            "gross_salary": round(gross_salary, 2),
            "total_deductions": round(total_deductions, 2),
            "net_salary": round(net_salary, 2),
            "fixed_salary": employee.get("fixed_salary", False),
            "salary_per_minute": round(salary_per_minute, 2)
        })
    
    total_gross_calc = round(sum([r["gross_salary"] for r in detailed_records]), 2)
    today_total_earnings_rounded = round(today_total_earnings, 2)
    print(f"DEBUG LIVE PAYROLL: Calculated total_gross={total_gross_calc}, today_total_earnings={today_total_earnings_rounded} from {len(detailed_records)} records")
    
    return {
        "month": current_month,
        "timestamp": now.isoformat(),
        "employees": detailed_records,
        "total_gross": total_gross_calc,
        "total_net": round(sum([r["net_salary"] for r in detailed_records]), 2),
        "total_deductions": round(sum([r["total_deductions"] for r in detailed_records]), 2),
        "total_allowances": round(sum([r["allowances"] for r in detailed_records]), 2),
        "today_total_earnings": today_total_earnings_rounded
    }


# ============= UTILITY FUNCTIONS =============
def capitalize_name(name: str) -> str:
    """Capitalize first letter of each word in a name"""
    return ' '.join(word.capitalize() for word in name.split())

def calculate_working_days(year: int, month: int, holidays: List[dict], saturday_enabled: bool = True, saturday_type: str = "full") -> dict:
    """
    Calculate working days for a given month considering:
    - Sundays (weekly off)
    - Public holidays from holiday calendar
    - Saturday settings (full day, half day, or off)
    """
    import calendar
    from datetime import date
    
    # Get total days in month
    total_days = calendar.monthrange(year, month)[1]
    
    # Count working days
    working_days = 0
    half_days = 0
    
    # Convert holidays to date strings for comparison
    holiday_dates = set()
    for holiday in holidays:
        try:
            holiday_date = datetime.fromisoformat(holiday['date']).date()
            if holiday_date.year == year and holiday_date.month == month:
                holiday_dates.add(holiday['date'])
        except:
            continue
    
    for day in range(1, total_days + 1):
        current_date = date(year, month, day)
        date_str = current_date.isoformat()
        weekday = current_date.weekday()  # 0=Monday, 6=Sunday
        
        # Skip Sundays
        if weekday == 6:
            continue
        
        # Skip holidays
        if date_str in holiday_dates:
            continue
        
        # Handle Saturday
        if weekday == 5:  # Saturday
            if not saturday_enabled:
                continue
            elif saturday_type == "half":
                half_days += 1
            else:  # full day
                working_days += 1
        else:
            working_days += 1
    
    # Convert half days to working days (2 half days = 1 full day)
    total_working_days = working_days + (half_days * 0.5)
    
    return {
        "total_days": total_days,
        "working_days": round(total_working_days, 1),
        "full_days": working_days,
        "half_days": half_days,
        "holidays": len(holiday_dates),
        "sundays": sum(1 for day in range(1, total_days + 1) if date(year, month, day).weekday() == 6)
    }

# ============= SETTINGS ENDPOINTS =============
@api_router.get("/settings")
async def get_settings(current_user: User = Depends(get_current_user)):
    if current_user.role == "super_admin":
        raise HTTPException(status_code=400, detail="Not applicable for super admin")
    
    settings = await db.settings.find_one({"company_id": current_user.company_id}, {"_id": 0})
    
    if not settings:
        # Create default settings
        default_settings = CompanySettings(company_id=current_user.company_id)
        await db.settings.insert_one(default_settings.model_dump())
        return default_settings
    
    return CompanySettings(**settings)

@api_router.put("/settings")
async def update_settings(updates: SettingsUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Create detailed log of changes
    settings_changes = ', '.join([f'{k}: {v}' for k, v in update_data.items() if k not in ['updated_at', '_id']])
    
    result = await db.settings.update_one(
        {"company_id": current_user.company_id},
        {"$set": update_data},
        upsert=True
    )
    
    # Log activity regardless of whether it was an insert or update
    await log_activity(current_user.company_id, current_user.id, current_user.name, "UPDATE_SETTINGS", f"Updated settings: {settings_changes}")
    
    return {"message": "Settings updated successfully"}

@api_router.post("/settings/holidays")
async def add_holiday(holiday: Holiday, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    result = await db.settings.update_one(
        {"company_id": current_user.company_id},
        {"$push": {"holidays": holiday.model_dump()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Settings not found")
    
    await log_activity(current_user.company_id, current_user.id, current_user.name, "ADD_HOLIDAY", f"Added holiday: {holiday.name} on {holiday.date}, Type: {holiday.type}")
    
    return {"message": "Holiday added successfully"}

@api_router.delete("/settings/holidays/{date}")
async def delete_holiday(date: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    result = await db.settings.update_one(
        {"company_id": current_user.company_id},
        {"$pull": {"holidays": {"date": date}}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Holiday not found")
    
    holiday_name = next((h['name'] for h in holidays if h['date'] == date), 'Unknown')
    await log_activity(current_user.company_id, current_user.id, current_user.name, "DELETE_HOLIDAY", f"Removed holiday: {holiday_name} on {date}")
    
    return {"message": "Holiday removed successfully"}

@api_router.get("/settings/working-days/{year}/{month}")
async def get_working_days(year: int, month: int, current_user: User = Depends(get_current_user)):
    if current_user.role == "super_admin":
        raise HTTPException(status_code=400, detail="Not applicable for super admin")
    
    # Get company settings
    settings = await db.settings.find_one({"company_id": current_user.company_id}, {"_id": 0})
    
    if not settings:
        # Use default settings
        holidays = []
        saturday_enabled = True
        saturday_type = "full"
    else:
        holidays = settings.get("holidays", [])
        saturday_enabled = settings.get("saturday_enabled", True)
        saturday_type = settings.get("saturday_type", "full")
    
    # Calculate working days
    result = calculate_working_days(year, month, holidays, saturday_enabled, saturday_type)
    
    return result

# ============= BRANDING ENDPOINTS =============
@api_router.post("/company/branding")
async def upload_branding(file: UploadFile = File(...), type: str = Form(...), current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    if type not in ["logo", "favicon"]:
        raise HTTPException(status_code=400, detail="Invalid type. Must be 'logo' or 'favicon'")
    
    try:
        # Read file and convert to base64
        contents = await file.read()
        base64_image = base64.b64encode(contents).decode('utf-8')
        data_url = f"data:{file.content_type};base64,{base64_image}"
        
        # Update settings with the uploaded image
        field_name = "company_logo" if type == "logo" else "favicon"
        await db.settings.update_one(
            {"company_id": current_user.company_id},
            {"$set": {field_name: data_url}},
            upsert=True
        )
        
        await log_activity(current_user.company_id, current_user.id, current_user.name, f"UPLOAD_{type.upper()}", f"Uploaded company {type}, File: {file.filename}, Size: {len(contents)} bytes, Type: {file.content_type}")
        
        return {"message": f"{type.capitalize()} uploaded successfully", field_name: data_url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/superadmin/branding")
async def upload_superadmin_branding(
    file: UploadFile = File(...), 
    type: str = Form(...), 
    company_id: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    
    if type not in ["logo", "favicon"]:
        raise HTTPException(status_code=400, detail="Invalid type. Must be 'logo' or 'favicon'")
    
    try:
        # Verify company exists
        company = await db.companies.find_one({"id": company_id})
        if not company:
            raise HTTPException(status_code=404, detail="Company not found")
        
        # Read file and convert to base64
        contents = await file.read()
        base64_image = base64.b64encode(contents).decode('utf-8')
        data_url = f"data:{file.content_type};base64,{base64_image}"
        
        # Update settings with the uploaded image
        field_name = "company_logo" if type == "logo" else "favicon"
        await db.settings.update_one(
            {"company_id": company_id},
            {"$set": {field_name: data_url}},
            upsert=True
        )
        
        await log_activity("SUPER_ADMIN", current_user.id, current_user.name, f"UPLOAD_{type.upper()}", f"Uploaded {type} for company {company['name']}")
        
        return {"message": f"{type.capitalize()} uploaded successfully for {company['name']}", field_name: data_url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============= PROFILE PICTURE ENDPOINTS =============
@api_router.post("/employees/profile-picture")
async def upload_employee_profile_pic(
    file: UploadFile = File(...),
    employee_id: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    try:
        # Read file and convert to base64
        contents = await file.read()
        base64_image = base64.b64encode(contents).decode('utf-8')
        data_url = f"data:{file.content_type};base64,{base64_image}"
        
        # Update employee profile picture
        await db.users.update_one(
            {"id": employee_id, "company_id": current_user.company_id},
            {"$set": {"profile_pic": data_url}}
        )
        
        return {"message": "Profile picture updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to upload profile picture: {str(e)}")


# ============= DEVICE ATTENDANCE IMPORT ENDPOINTS =============

@api_router.post("/attendance/parse-device-import")
async def parse_device_import(request: DeviceImportParseRequest, current_user: User = Depends(get_current_user)):
    """Parse attendance device file - Direct Python parsing"""
    
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    print(f"DEBUG: Parsing device import, file length: {len(request.file_content)}")
    
    try:
        import json
        from collections import defaultdict
        import re
        
        # Check if it's an Excel file
        if request.file_content.startswith('[EXCEL_FILE]'):
            print("DEBUG: Processing Excel file with AI")
            # Extract the base64 data
            parts = request.file_content.split('\n', 2)
            if len(parts) < 3:
                raise HTTPException(status_code=400, detail="Invalid Excel file format")
            
            filename = parts[1]
            base64_data = parts[2].split(',', 1)[1] if ',' in parts[2] else parts[2]
            
            # Decode base64 and save to temp file
            import base64
            import io
            import os
            import tempfile
            import json
            from openpyxl import load_workbook
            from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType
            from dotenv import load_dotenv
            
            load_dotenv()
            
            excel_data = base64.b64decode(base64_data)
            
            # Save to temporary file for AI processing
            with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(excel_data)
                temp_file_path = temp_file.name
            
            try:
                # Also convert Excel to readable text format for better AI understanding
                wb = load_workbook(io.BytesIO(excel_data))
                sheet = wb.active
                
                # Get first 30 rows to show structure to AI
                excel_text = "EXCEL FILE STRUCTURE:\n"
                excel_text += "=" * 80 + "\n"
                for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if i > 30:  # Show first 30 rows for context
                        excel_text += f"\n... (showing first 30 rows of {sheet.max_row} total rows)\n"
                        break
                    if any(cell is not None for cell in row):
                        excel_text += f"Row {i}: {row}\n"
                
                # Initialize AI chat with Emergent LLM Key
                api_key = os.getenv('EMERGENT_LLM_KEY', '')
                if not api_key:
                    raise HTTPException(status_code=500, detail="LLM API key not configured")
                
                chat = LlmChat(
                    api_key=api_key,
                    session_id=f"attendance-parse-{request.company_id}",
                    system_message="""You are an expert at analyzing fingerprint attendance device Excel files. 
Your task is to extract employee attendance data (punch in/out times) from any Excel format, regardless of how it's structured.

IMPORTANT INSTRUCTIONS:
1. Identify the employee ID column (might be called: Enroll No, Employee ID, Badge ID, etc.)
2. Identify date columns (could be in headers or in data rows)
3. Identify time/punch data (IN/OUT times, or just timestamps)
4. Extract ALL unique employee IDs
5. Extract ALL punch records with: employee_id, date, time, type (in/out)

Return your response as a valid JSON object with this EXACT structure:
{
  "unique_vendor_ids": ["id1", "id2", "id3", ...],
  "records": [
    {
      "vendor_id": "employee_id",
      "datetime": "YYYY-MM-DD HH:MM",
      "date": "YYYY-MM-DD",
      "time": "HH:MM",
      "record_type": "punch_in" or "punch_out"
    }
  ],
  "format_detected": "description of the format you detected"
}

CRITICAL: Ensure dates are in YYYY-MM-DD format and times are in HH:MM format.
Extract BOTH punch_in and punch_out records when available.
If only timestamps are available (no explicit IN/OUT), alternate between punch_in and punch_out."""
                ).with_model("gemini", "gemini-2.0-flash")
                
                # Create file attachment for AI to analyze
                excel_file_obj = FileContentWithMimeType(
                    file_path=temp_file_path,
                    mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Send message with file and context
                user_message = UserMessage(
                    text=f"""Analyze this attendance Excel file and extract ALL employee punch records.

File name: {filename}

Here's the structure I can see:
{excel_text}

Please extract:
1. All unique employee/device IDs
2. All punch records (IN and OUT times)
3. Convert all dates to YYYY-MM-DD format
4. Convert all times to HH:MM format

Return ONLY the JSON response, no additional text.""",
                    file_contents=[excel_file_obj]
                )
                
                print("DEBUG: Sending Excel file to AI for analysis...")
                response_text = await chat.send_message(user_message)
                print(f"DEBUG: AI response received: {response_text[:500]}...")
                
                # Parse AI response
                # Remove markdown code blocks if present
                response_text = response_text.strip()
                if response_text.startswith('```'):
                    # Extract JSON from markdown code block
                    lines = response_text.split('\n')
                    response_text = '\n'.join(lines[1:-1])
                
                # Try to find JSON in the response
                import re
                json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(0)
                
                ai_result = json.loads(response_text)
                
                # Validate and process AI response
                records = ai_result.get('records', [])
                unique_vendor_ids = ai_result.get('unique_vendor_ids', [])
                
                # Extract dates for date range
                dates = set()
                for record in records:
                    if 'date' in record:
                        dates.add(record['date'])
                
                sorted_dates = sorted(list(dates))
                date_range = {
                    "start": sorted_dates[0] if sorted_dates else None,
                    "end": sorted_dates[-1] if sorted_dates else None
                }
                
                parsed_data = {
                    "format_detected": ai_result.get('format_detected', 'AI-detected format'),
                    "records": records,
                    "unique_vendor_ids": sorted(unique_vendor_ids),
                    "date_range": date_range,
                    "total_records": len(records)
                }
                
                print(f"DEBUG: AI parsed {len(records)} records for {len(unique_vendor_ids)} employees")
                
                await log_activity(
                    request.company_id,
                    current_user.id,
                    current_user.name,
                    "PARSE_DEVICE_IMPORT",
                    f"AI parsed Excel import: {len(records)} records found"
                )
                
                return {
                    "success": True,
                    "data": parsed_data
                }
                
            except json.JSONDecodeError as e:
                print(f"ERROR: Failed to parse AI response as JSON: {e}")
                print(f"AI Response: {response_text}")
                raise HTTPException(status_code=500, detail=f"AI returned invalid JSON: {str(e)}")
            except Exception as e:
                print(f"ERROR: AI parsing failed: {e}")
                import traceback
                traceback.print_exc()
                raise HTTPException(status_code=500, detail=f"AI parsing error: {str(e)}")
            finally:
                # Clean up temp file
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
        
        # Parse the file content
        lines = request.file_content.strip().split('\n')
        records = []
        unique_vendor_ids = set()
        dates = set()
        
        print(f"DEBUG: Processing {len(lines)} lines")
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Split by tabs or multiple spaces
            parts = re.split(r'\t+|\s{2,}', line)
            
            if len(parts) >= 2:
                vendor_id = parts[0].strip()
                datetime_str = parts[1].strip()
                
                # Parse datetime
                try:
                    # Try to extract date and time
                    if ' ' in datetime_str:
                        date_part, time_part = datetime_str.split(' ', 1)
                        
                        records.append({
                            "vendor_id": vendor_id,
                            "datetime": datetime_str,
                            "date": date_part,
                            "time": time_part,
                            "record_type": "punch"
                        })
                        
                        unique_vendor_ids.add(vendor_id)
                        dates.add(date_part)
                except Exception as e:
                    print(f"DEBUG: Failed to parse line: {line}, error: {e}")
                    continue
        
        # Sort dates to get range
        sorted_dates = sorted(list(dates))
        date_range = {
            "start": sorted_dates[0] if sorted_dates else None,
            "end": sorted_dates[-1] if sorted_dates else None
        }
        
        # Create response
        parsed_data = {
            "format_detected": "Tab or space-separated format with vendor_id and datetime",
            "records": records,
            "unique_vendor_ids": sorted(list(unique_vendor_ids)),
            "date_range": date_range,
            "total_records": len(records)
        }
        
        print(f"DEBUG: Parsed {len(records)} records, {len(unique_vendor_ids)} unique IDs")
        
        await log_activity(
            request.company_id,
            current_user.id,
            current_user.name,
            "PARSE_DEVICE_IMPORT",
            f"Parsed device import file: {len(records)} records found"
        )
        
        return {
            "success": True,
            "data": parsed_data
        }
        
    except Exception as e:
        print(f"Parse Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to parse file: {str(e)}")


@api_router.post("/attendance/import-device-data")
async def import_device_data(request: DeviceImportRequest, current_user: User = Depends(get_current_user)):
    """Import device attendance data with ID mapping"""
    
    if current_user.role not in ["admin", "manager", "accountant"]:
        raise HTTPException(status_code=403, detail="Admin, manager or accountant access required")
    
    # Create mapping dictionary
    id_mapping = {mapping.vendor_id: mapping.employee_id for mapping in request.mappings}
    
    imported_count = 0
    skipped_count = 0
    overwritten_count = 0
    errors = []
    
    # Group records by vendor_id and date
    from collections import defaultdict
    grouped_records = defaultdict(list)
    
    for record in request.parsed_records:
        vendor_id = record.vendor_id
        if vendor_id in id_mapping:
            employee_id = id_mapping[vendor_id]
            grouped_records[(employee_id, record.date)].append(record)
    
    # Process each employee-date group
    for (employee_id, date), records in grouped_records.items():
        try:
            # Get employee
            employee = await db.users.find_one({"id": employee_id, "company_id": request.company_id})
            if not employee:
                errors.append(f"Employee {employee_id} not found for date {date}")
                continue
            
            # Sort records by time to get check-in and check-out
            sorted_records = sorted(records, key=lambda r: r.time)
            check_in_time = sorted_records[0].time
            check_out_time = sorted_records[-1].time if len(sorted_records) > 1 else None
            
            # Check if attendance already exists
            existing = await db.attendance.find_one({
                "company_id": request.company_id,
                "employee_id": employee_id,
                "date": date
            })
            
            if existing:
                if request.duplicate_action == "skip":
                    skipped_count += 1
                    continue
                elif request.duplicate_action == "overwrite":
                    # Update existing record
                    await db.attendance.update_one(
                        {"id": existing["id"]},
                        {"$set": {
                            "check_in": f"{date}T{check_in_time}",
                            "check_out": f"{date}T{check_out_time}" if check_out_time else None,
                            "status": "present",
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                            "updated_by": current_user.id
                        }}
                    )
                    overwritten_count += 1
                    continue
            
            # Create new attendance record
            new_attendance = {
                "id": str(uuid.uuid4()),
                "company_id": request.company_id,
                "employee_id": employee_id,
                "employee_name": capitalize_name(employee["name"]),
                "date": date,
                "check_in": f"{date}T{check_in_time}",
                "check_out": f"{date}T{check_out_time}" if check_out_time else None,
                "status": "present",
                "created_by": current_user.id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.attendance.insert_one(new_attendance)
            imported_count += 1
            
        except Exception as e:
            errors.append(f"Error processing {employee_id} on {date}: {str(e)}")
    
    await log_activity(
        request.company_id,
        current_user.id,
        current_user.name,
        "IMPORT_DEVICE_ATTENDANCE",
        f"Imported {imported_count} records, skipped {skipped_count}, overwritten {overwritten_count}"
    )
    
    return {
        "success": True,
        "imported": imported_count,
        "skipped": skipped_count,
        "overwritten": overwritten_count,
        "errors": errors
    }

@api_router.post("/upload/profile-pic")
async def upload_profile_pic(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    try:
        # Read file and convert to base64
        contents = await file.read()
        base64_image = base64.b64encode(contents).decode('utf-8')
        data_url = f"data:{file.content_type};base64,{base64_image}"
        
        # Update user profile picture
        await db.users.update_one(
            {"id": current_user.id},
            {"$set": {"profile_pic": data_url}}
        )
        
        await log_activity(current_user.company_id or "SUPER_ADMIN", current_user.id, current_user.name, "UPDATE_PROFILE_PIC", "Updated profile picture")
        
        return {"message": "Profile picture uploaded successfully", "profile_pic": data_url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============= LOCATION TRACKING ENDPOINTS =============

@api_router.post("/location/tracking/start")
async def start_location_tracking(current_user: User = Depends(get_current_user)):
    """Start a new location tracking session for the current user"""
    
    # Check if there's already an active tracking session
    existing_session = await db.tracking_sessions.find_one({
        "company_id": current_user.company_id,
        "employee_id": current_user.employee_id or current_user.id,
        "status": "active"
    })
    
    if existing_session:
        return {
            "message": "Tracking session already active",
            "session_id": existing_session["id"],
            "start_time": existing_session["start_time"]
        }
    
    # Create new tracking session
    session = {
        "id": str(uuid.uuid4()),
        "company_id": current_user.company_id,
        "employee_id": current_user.employee_id or current_user.id,
        "employee_name": capitalize_name(current_user.name),
        "start_time": datetime.now(timezone.utc).isoformat(),
        "end_time": None,
        "status": "active",
        "locations": [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.tracking_sessions.insert_one(session)
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "START_LOCATION_TRACKING",
        f"Started location tracking session"
    )
    
    return {
        "message": "Location tracking started",
        "session_id": session["id"],
        "start_time": session["start_time"]
    }

@api_router.post("/location/tracking/update")
async def update_location(location_data: LocationUpdate, current_user: User = Depends(get_current_user)):
    """Add a location point to an active tracking session"""
    
    # Verify session exists and is active
    session = await db.tracking_sessions.find_one({
        "id": location_data.session_id,
        "company_id": current_user.company_id,
        "employee_id": current_user.employee_id or current_user.id,
        "status": "active"
    })
    
    if not session:
        raise HTTPException(status_code=404, detail="Active tracking session not found")
    
    # Create location point
    location_point = {
        "latitude": location_data.latitude,
        "longitude": location_data.longitude,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "accuracy": location_data.accuracy
    }
    
    # Add location to session
    await db.tracking_sessions.update_one(
        {"id": location_data.session_id},
        {"$push": {"locations": location_point}}
    )
    
    return {
        "message": "Location updated",
        "timestamp": location_point["timestamp"]
    }

@api_router.post("/location/tracking/stop")
async def stop_location_tracking(session_data: dict, current_user: User = Depends(get_current_user)):
    """Stop an active location tracking session"""
    
    session_id = session_data.get("session_id")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id is required")
    
    # Verify session exists and is active
    session = await db.tracking_sessions.find_one({
        "id": session_id,
        "company_id": current_user.company_id,
        "employee_id": current_user.employee_id or current_user.id,
        "status": "active"
    })
    
    if not session:
        raise HTTPException(status_code=404, detail="Active tracking session not found")
    
    # Stop the session
    end_time = datetime.now(timezone.utc).isoformat()
    await db.tracking_sessions.update_one(
        {"id": session_id},
        {
            "$set": {
                "status": "stopped",
                "end_time": end_time
            }
        }
    )
    
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "STOP_LOCATION_TRACKING",
        f"Stopped location tracking session (Duration: {len(session.get('locations', []))} points)"
    )
    
    return {
        "message": "Location tracking stopped",
        "session_id": session_id,
        "end_time": end_time,
        "total_locations": len(session.get("locations", []))
    }

@api_router.post("/attendance/mark-with-location")
async def mark_attendance_with_location(attendance_data: AttendanceWithLocation, current_user: User = Depends(get_current_user)):
    """Mark attendance with location snapshot"""
    
    # Validate employee
    employee_id = attendance_data.employee_id or current_user.employee_id or current_user.id
    
    # Debug logging
    print(f"DEBUG: Trying to find employee - employee_id: {employee_id}, company_id: {current_user.company_id}")
    print(f"DEBUG: current_user.id: {current_user.id}, current_user.employee_id: {current_user.employee_id}")
    
    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    
    if not employee:
        # Try alternative lookup - maybe the user doesn't have employee_id field
        employee = await db.users.find_one({"id": current_user.id})
        print(f"DEBUG: Alternative lookup result: {employee}")
        
        if not employee:
            raise HTTPException(
                status_code=404, 
                detail=f"Employee not found. Searched for id={employee_id}, company_id={current_user.company_id}"
            )
    
    # Check how many attendance records exist for this date (allow up to 10 per day)
    existing_count = await db.attendance.count_documents({
        "company_id": current_user.company_id,
        "employee_id": employee_id,
        "date": attendance_data.date
    })
    
    if existing_count >= 10:
        raise HTTPException(
            status_code=400, 
            detail=f"Daily limit exceeded. Maximum 10 attendance records per day allowed. Current count: {existing_count}"
        )
    
    # Create attendance record with location
    check_in_datetime = None
    check_out_datetime = None
    
    if attendance_data.check_in:
        check_in_datetime = f"{attendance_data.date}T{attendance_data.check_in}:00"
    else:
        # Use current time as check-in if not provided
        current_time = datetime.now(timezone.utc).strftime("%H:%M")
        check_in_datetime = f"{attendance_data.date}T{current_time}:00"
    
    if attendance_data.check_out:
        check_out_datetime = f"{attendance_data.date}T{attendance_data.check_out}:00"
    
    new_attendance = {
        "id": str(uuid.uuid4()),
        "company_id": current_user.company_id,
        "employee_id": employee_id,
        "employee_name": capitalize_name(employee["name"]),
        "date": attendance_data.date,
        "check_in": check_in_datetime,
        "check_out": check_out_datetime,
        "status": attendance_data.status,
        "leave_type": attendance_data.leave_type,
        "location": {
            "latitude": attendance_data.latitude,
            "longitude": attendance_data.longitude,
            "accuracy": attendance_data.accuracy,
            "address": attendance_data.address,
            "map_snapshot": attendance_data.map_snapshot,
            "captured_at": datetime.now(timezone.utc).isoformat()
        },
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    attendance_response = new_attendance.copy()
    await db.attendance.insert_one(new_attendance)
    
    await log_activity(
        current_user.company_id,
        current_user.id,
        current_user.name,
        "ADD_ATTENDANCE_WITH_LOCATION",
        f"Marked attendance with location for {capitalize_name(employee['name'])} on {attendance_data.date} at ({attendance_data.latitude:.4f}, {attendance_data.longitude:.4f})"
    )
    
    return {"message": "Attendance marked with location", "attendance": attendance_response}

@api_router.get("/attendance/fingerprint/{company_short_code}/{fingerprint_id}")
async def mark_attendance_by_fingerprint(company_short_code: str, fingerprint_id: str):
    """
    Mark attendance using company short code and fingerprint ID (no authentication required)
    - company_short_code: Company identifier (max 20 chars)
    - fingerprint_id: Employee fingerprint ID
    - If no attendance for today: Mark check-in
    - If attendance exists without check-out and >10 minutes passed: Mark check-out
    """
    
    # Validate company short code
    if not company_short_code or company_short_code.strip() == '':
        return {"success": False, "message": "Missing company short code"}
    
    # Find company by short code
    company = await db.companies.find_one({"short_code": company_short_code}, {"_id": 0})
    
    if not company:
        return {"success": False, "message": "Invalid company short code"}
    
    # Find user by fingerprint_id within this specific company
    user = await db.users.find_one({
        "fingerprint_id": fingerprint_id,
        "company_id": company["id"]
    }, {"_id": 0})
    
    if not user:
        return {"success": False, "message": "No User"}
    
    # Get current time in Sri Lanka timezone (UTC+5:30)
    sri_lanka_tz = pytz.timezone('Asia/Colombo')
    current_time_lk = datetime.now(sri_lanka_tz)
    today = current_time_lk.strftime("%Y-%m-%d")
    current_time_str = current_time_lk.strftime("%H:%M")
    
    # For time comparison, use UTC
    current_time_utc = datetime.now(timezone.utc)
    
    # Check if attendance record exists for today
    attendance = await db.attendance.find_one({
        "company_id": user.get("company_id"),
        "employee_id": user["id"],
        "date": today
    }, {"_id": 0})
    
    if not attendance:
        # No attendance for today - Mark check-in
        check_in_datetime = f"{today}T{current_time_str}:00"
        
        new_attendance = {
            "id": str(uuid.uuid4()),
            "company_id": user.get("company_id"),
            "employee_id": user["id"],
            "employee_name": capitalize_name(user["name"]),
            "date": today,
            "check_in": check_in_datetime,
            "check_out": None,
            "status": "present",
            "leave_type": "",
            "created_by": user["id"],  # Self-marked via fingerprint
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.attendance.insert_one(new_attendance)
        
        return {
            "success": True,
            "message": f"Attendance Success - {capitalize_name(user['name'])}",
            "action": "check_in",
            "time": current_time_str
        }
    
    else:
        # Attendance exists - Check if we should mark check-out
        if attendance.get("check_out"):
            # Already has check-out
            return {
                "success": False,
                "message": f"Attendance already completed for {capitalize_name(user['name'])} today"
            }
        
        # Parse check-in time to verify 10-minute difference
        check_in_str = attendance.get("check_in", "")
        if check_in_str:
            try:
                # Parse check_in datetime (format: "2025-12-26T09:30:00")
                # The stored time is in Sri Lanka timezone (naive datetime)
                check_in_dt = datetime.fromisoformat(check_in_str)
                
                # Make it timezone-aware in Sri Lanka timezone
                if check_in_dt.tzinfo is None:
                    check_in_dt = sri_lanka_tz.localize(check_in_dt)
                
                # Calculate time difference using Sri Lanka timezone
                time_diff = (current_time_lk - check_in_dt).total_seconds() / 60  # in minutes
                
                if time_diff < 10:
                    return {
                        "success": False,
                        "message": f"Please wait {int(10 - time_diff)} more minutes before marking leaving"
                    }
                
            except Exception as e:
                print(f"Error parsing check_in time: {e}")
                # Continue to mark check-out even if parsing fails
        
        # Mark check-out
        check_out_datetime = f"{today}T{current_time_str}:00"
        
        await db.attendance.update_one(
            {"id": attendance["id"]},
            {"$set": {"check_out": check_out_datetime}}
        )
        
        return {
            "success": True,
            "message": f"Leaving Marked Success - {capitalize_name(user['name'])}",
            "action": "check_out",
            "time": current_time_str
        }

@api_router.get("/location/tracking/history")
async def get_tracking_history(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get location tracking history for current user"""
    
    employee_id = current_user.employee_id or current_user.id
    
    # Build query
    query = {
        "company_id": current_user.company_id,
        "employee_id": employee_id
    }
    
    # Add date filtering if provided (start_time is ISO datetime)
    if from_date or to_date:
        date_filter = {}
        if from_date:
            date_filter["$gte"] = f"{from_date}T00:00:00"
        if to_date:
            date_filter["$lte"] = f"{to_date}T23:59:59"
        query["start_time"] = date_filter
    
    # Get tracking sessions
    sessions = await db.tracking_sessions.find(
        query,
        {"_id": 0}
    ).sort("start_time", -1).to_list(length=None)
    
    return {"sessions": sessions, "total": len(sessions)}

@api_router.get("/location/reports/employee/{employee_id}")
async def get_employee_location_report(
    employee_id: str,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Admin: Get location reports for a specific employee"""
    
    if current_user.role not in ["admin", "manager", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin or manager access required")
    
    # Verify employee exists (check both with company_id and without for super admins)
    employee = await db.users.find_one({"id": employee_id, "company_id": current_user.company_id})
    if not employee:
        # Try finding user without company_id restriction (for super admin testing)
        employee = await db.users.find_one({"id": employee_id})
    
    # If still not found, check if this employee_id has tracking data and create placeholder
    if not employee:
        # Check if there's tracking data for this employee_id
        has_tracking = await db.tracking_sessions.find_one({"employee_id": employee_id, "company_id": current_user.company_id})
        has_attendance = await db.attendance.find_one({"employee_id": employee_id, "company_id": current_user.company_id})
        
        if not (has_tracking or has_attendance):
            raise HTTPException(status_code=404, detail="Employee not found")
        
        # Create placeholder employee from tracking session data
        session_with_name = has_tracking or has_attendance
        employee = {
            "id": employee_id,
            "name": session_with_name.get("employee_name", "Unknown User"),
            "mobile": "N/A",
            "position": "Super Admin" if employee_id == "SUPER-ADMIN" else "Unknown"
        }
    
    # Build query
    query = {
        "company_id": current_user.company_id,
        "employee_id": employee_id
    }
    
    # Add date filtering if provided (start_time is ISO datetime)
    if from_date or to_date:
        tracking_date_filter = {}
        if from_date:
            tracking_date_filter["$gte"] = f"{from_date}T00:00:00"
        if to_date:
            tracking_date_filter["$lte"] = f"{to_date}T23:59:59"
        query["start_time"] = tracking_date_filter
    
    # Get tracking sessions
    tracking_sessions = await db.tracking_sessions.find(
        query,
        {"_id": 0}
    ).sort("start_time", -1).to_list(length=None)
    
    # Get attendance records with location
    attendance_query = {
        "company_id": current_user.company_id,
        "employee_id": employee_id,
        "location": {"$exists": True}
    }
    
    # Add date filtering for attendance (date is YYYY-MM-DD)
    if from_date or to_date:
        attendance_date_filter = {}
        if from_date:
            attendance_date_filter["$gte"] = from_date
        if to_date:
            attendance_date_filter["$lte"] = to_date
        attendance_query["date"] = attendance_date_filter
    
    attendance_records = await db.attendance.find(
        attendance_query,
        {"_id": 0}
    ).sort("date", -1).to_list(length=None)
    
    return {
        "employee": {
            "id": employee["id"],
            "name": capitalize_name(employee["name"]),
            "mobile": employee["mobile"],
            "position": employee.get("position")
        },
        "tracking_sessions": tracking_sessions,
        "attendance_with_location": attendance_records,
        "summary": {
            "total_tracking_sessions": len(tracking_sessions),
            "total_attendance_with_location": len(attendance_records),
            "total_location_points": sum(len(session.get("locations", [])) for session in tracking_sessions)
        }
    }

@api_router.get("/location/reports/all")
async def get_all_location_reports(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Admin: Get location reports for all employees"""
    
    if current_user.role not in ["admin", "manager", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin or manager access required")
    
    # Build date filter for tracking (start_time is ISO datetime string)
    tracking_date_filter = {}
    if from_date or to_date:
        if from_date:
            tracking_date_filter["$gte"] = f"{from_date}T00:00:00"
        if to_date:
            tracking_date_filter["$lte"] = f"{to_date}T23:59:59"
    
    # Get all tracking sessions for the company
    tracking_query = {"company_id": current_user.company_id}
    if tracking_date_filter:
        tracking_query["start_time"] = tracking_date_filter
    
    all_tracking_sessions = await db.tracking_sessions.find(
        tracking_query,
        {"_id": 0}
    ).sort("start_time", -1).to_list(length=None)
    
    # Build date filter for attendance (date is YYYY-MM-DD string)
    attendance_date_filter = {}
    if from_date or to_date:
        if from_date:
            attendance_date_filter["$gte"] = from_date
        if to_date:
            attendance_date_filter["$lte"] = to_date
    
    # Get all attendance with location for the company
    attendance_query = {
        "company_id": current_user.company_id,
        "location": {"$exists": True}
    }
    if attendance_date_filter:
        attendance_query["date"] = attendance_date_filter
    
    all_attendance = await db.attendance.find(
        attendance_query,
        {"_id": 0}
    ).sort("date", -1).to_list(length=None)
    
    # Extract unique employee_ids from tracking sessions and attendance
    employee_ids = set()
    for session in all_tracking_sessions:
        if session.get("employee_id"):
            employee_ids.add(session["employee_id"])
    for att in all_attendance:
        if att.get("employee_id"):
            employee_ids.add(att["employee_id"])
    
    print(f"DEBUG: Unique employee_ids with location data: {employee_ids}")
    
    # Fetch user information for all these employee_ids
    users_with_data = await db.users.find(
        {"id": {"$in": list(employee_ids)}},
        {"_id": 0, "id": 1, "name": 1, "mobile": 1, "position": 1, "role": 1}
    ).to_list(length=None)
    
    print(f"DEBUG: Found {len(users_with_data)} users from database")
    
    # For any employee_ids not found in users, create placeholder from session data
    found_ids = {user["id"] for user in users_with_data}
    missing_ids = employee_ids - found_ids
    
    if missing_ids:
        print(f"DEBUG: Missing user IDs: {missing_ids}")
        # Add placeholder users from tracking session data
        for emp_id in missing_ids:
            # Find a session with this employee_id to get the name
            session_with_name = next((s for s in all_tracking_sessions if s.get("employee_id") == emp_id), None)
            if session_with_name:
                users_with_data.append({
                    "id": emp_id,
                    "name": session_with_name.get("employee_name", "Unknown User"),
                    "mobile": "N/A",
                    "position": "Super Admin" if emp_id == "SUPER-ADMIN" else "Unknown",
                    "role": "super_admin" if emp_id == "SUPER-ADMIN" else "unknown"
                })
    
    print(f"DEBUG: Total users with location data (including placeholders): {len(users_with_data)}")
    print(f"DEBUG: Total tracking sessions: {len(all_tracking_sessions)}")
    print(f"DEBUG: Total attendance records: {len(all_attendance)}")
    
    # Group by employee
    employee_reports = []
    
    for user in users_with_data:
        # Match by employee_id
        emp_tracking = [s for s in all_tracking_sessions if s.get("employee_id") == user["id"]]
        emp_attendance = [a for a in all_attendance if a.get("employee_id") == user["id"]]
        
        print(f"DEBUG: User {user.get('name')} (id={user['id']}, role={user.get('role')}): {len(emp_tracking)} tracking, {len(emp_attendance)} attendance")
        
        if emp_tracking or emp_attendance:
            employee_reports.append({
                "employee": {
                    "id": user["id"],
                    "name": capitalize_name(user["name"]),
                    "mobile": user.get("mobile", "N/A"),
                    "position": user.get("position", user.get("role", "").title())
                },
                "tracking_sessions_count": len(emp_tracking),
                "attendance_with_location_count": len(emp_attendance),
                "total_location_points": sum(len(s.get("locations", [])) for s in emp_tracking),
                "latest_tracking": emp_tracking[0] if emp_tracking else None,
                "latest_attendance": emp_attendance[0] if emp_attendance else None
            })
    
    return {
        "employees": employee_reports,
        "summary": {
            "total_employees_with_data": len(employee_reports),
            "total_tracking_sessions": len(all_tracking_sessions),
            "total_attendance_with_location": len(all_attendance),
            "total_location_points": sum(len(s.get("locations", [])) for s in all_tracking_sessions)
        }
    }


# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
